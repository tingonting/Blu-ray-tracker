import random
import streamlit.components.v1 as components
import difflib
import re
import urllib.parse
import datetime
import concurrent.futures
import openpyxl
import pandas as pd
import streamlit as st

# Set up page configuration & layout
st.set_page_config(
    page_title="Conor's Blu-ray Hub", page_icon="🎬", layout="wide"
)

# Custom styling — "Blu-ray Hub" theme
# Palette drawn from the format itself: Blu-ray's 405nm laser gives the violet,
# a disc's iridescent shimmer gives the teal, and cinema marquee bulbs give the amber.
st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Bebas+Neue&family=Manrope:wght@400;600;800&family=JetBrains+Mono:wght@500;700&display=swap');

    :root {
        --ink: #0A0B14;
        --panel: #14162A;
        --violet: #6C5CE8;
        --teal: #2DD4BF;
        --amber: #F0A83B;
        --paper: #EDECF5;
        --mute: #8B889E;
    }

    /* Content width — narrow on phones (unchanged from before), progressively
       roomier on iPad and desktop instead of sitting in a fixed 600px column
       with dead space on either side. layout="wide" above removes Streamlit's
       own centered-mode cap so this is the only width rule in play. */
    .block-container {
        max-width: 600px;
        margin: 0 auto;
        padding-top: 2rem;
    }
    @media (min-width: 768px) {
        .block-container { max-width: 820px; }
    }
    @media (min-width: 1200px) {
        .block-container { max-width: 1000px; }
    }

    .stApp {
        font-family: 'Manrope', sans-serif;
    }

    /* Body text only — deliberately NOT applied to span/div broadly, since
       Streamlit renders icons (expander arrows, checkmarks, etc.) as icon-font
       ligatures inside spans, and overriding their font breaks those glyphs. */
    .stApp p, .stApp label, .stApp li, .stApp td, .stApp th {
        font-family: 'Manrope', sans-serif;
    }

    /* Headline / section-header treatment */
    h1, h2, h3 {
        font-family: 'Bebas Neue', sans-serif !important;
        letter-spacing: 0.04em;
        text-transform: uppercase;
    }

    /* App title: violet-to-teal gradient, like a disc catching the light */
    .app-title {
        text-align: center;
        font-family: 'Bebas Neue', sans-serif;
        font-size: 2.6rem;
        letter-spacing: 0.06em;
        margin-bottom: 0px;
        background: linear-gradient(100deg, var(--violet) 20%, var(--teal) 80%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
    }
    .app-subtitle {
        text-align: center;
        color: var(--mute);
        font-family: 'JetBrains Mono', monospace;
        font-size: 0.8rem;
        letter-spacing: 0.12em;
        text-transform: uppercase;
        margin-top: -4px;
    }

    /* Sprocket-hole divider — a nod to physical film reels, standing in for st.divider() */
    .film-divider {
        display: flex;
        justify-content: space-between;
        margin: 18px 2px;
        opacity: 0.55;
    }
    .film-divider span {
        width: 7px;
        height: 7px;
        border-radius: 2px;
        background: var(--violet);
    }
    .film-divider span:nth-child(3n) { background: var(--teal); }

    /* Card/Metric styling */
    div[data-testid="stMetric"] {
        background: linear-gradient(160deg, rgba(108, 92, 232, 0.12), rgba(45, 212, 191, 0.05));
        border: 1px solid rgba(108, 92, 232, 0.25);
        padding: 10px;
        border-radius: 14px;
        text-align: center;
    }
    div[data-testid="stMetricValue"] {
        font-family: 'JetBrains Mono', monospace;
        color: var(--amber);
    }

    /* Winner card for random picker — marquee spotlight */
    .winner-box {
        position: relative;
        overflow: hidden;
        background: linear-gradient(135deg, rgba(108, 92, 232, 0.18), rgba(45, 212, 191, 0.1));
        border: 2px solid var(--amber);
        padding: 20px;
        border-radius: 18px;
        text-align: center;
        margin-top: 15px;
        margin-bottom: 15px;
        animation: winner-glow 4s ease-in-out infinite;
    }

    /* A light sweep across the card, like a marquee catching the light */
    .winner-box::before {
        content: "";
        position: absolute;
        top: 0;
        left: -60%;
        width: 40%;
        height: 100%;
        background: linear-gradient(100deg, transparent, rgba(255, 255, 255, 0.16), transparent);
        animation: winner-shimmer 3.6s ease-in-out infinite;
        pointer-events: none;
    }

    @keyframes winner-shimmer {
        0%   { left: -60%; }
        55%  { left: 130%; }
        100% { left: 130%; }
    }

    @keyframes winner-glow {
        0%, 100% { box-shadow: 0 4px 20px rgba(240, 168, 59, 0.22); border-color: var(--amber); }
        50%      { box-shadow: 0 4px 24px rgba(108, 92, 232, 0.3); border-color: var(--violet); }
    }

    /* Format badges — Blu-ray in violet (the laser color), 4K UHD in teal (the shimmer color) */
    .format-badge {
        display: inline-block;
        padding: 1px 9px;
        border-radius: 20px;
        font-size: 0.85em;
        font-weight: 700;
    }
    .format-badge.bluray {
        background: rgba(108, 92, 232, 0.18);
        color: #A79CFF;
    }
    .format-badge.uhd4k {
        background: rgba(45, 212, 191, 0.18);
        color: var(--teal);
    }
    .format-badge.other {
        background: rgba(139, 136, 158, 0.18);
        color: var(--mute);
    }

    /* Make buttons touch-friendly and prominent */
    .stButton button {
        width: 100%;
        border-radius: 12px;
        font-weight: 700;
        padding: 0.6rem 1rem;
        font-family: 'Manrope', sans-serif;
    }

    /* Form container look */
    div[data-testid="stForm"] {
        border: 1px solid rgba(108, 92, 232, 0.2);
        border-radius: 16px;
        padding: 15px;
        background-color: rgba(108, 92, 232, 0.04);
    }

    /* Dataframe responsiveness */
    div[data-testid="stDataFrame"] {
        width: 100%;
        border-radius: 12px;
    }

    /* Tabs: quiet until selected, then lit up in violet with a teal underline */
    button[data-baseweb="tab"] {
        font-family: 'Manrope', sans-serif;
        font-weight: 600;
        color: var(--mute);
    }
    button[data-baseweb="tab"][aria-selected="true"] {
        color: var(--paper);
    }
    div[data-baseweb="tab-highlight"] {
        background-color: var(--teal) !important;
    }

    /* Rating stars -- consistent amber styling wherever a rating is
       displayed as text, rather than however the phone's own emoji font
       happens to render a bare star character. */
    .stars-display {
        color: var(--amber);
        letter-spacing: 0.08em;
        font-size: 1.05em;
        text-shadow: 0 0 8px rgba(240, 168, 59, 0.35);
    }

    /* Micro-interactions: buttons lift slightly on hover, settle on press,
       instead of the flat instant state-change of a default button. */
    .stButton button {
        transition: transform 0.15s ease, box-shadow 0.15s ease;
    }
    .stButton button:hover {
        transform: translateY(-1px);
        box-shadow: 0 4px 14px rgba(108, 92, 232, 0.35);
    }
    .stButton button:active {
        transform: translateY(0px) scale(0.97);
    }

    /* Themed loading spinner instead of Streamlit's default color */
    div[data-testid="stSpinner"] svg {
        color: var(--violet) !important;
    }

    /* Scrollbar, restyled to match the theme instead of the browser default */
    ::-webkit-scrollbar {
        width: 10px;
        height: 10px;
    }
    ::-webkit-scrollbar-track {
        background: var(--ink);
    }
    ::-webkit-scrollbar-thumb {
        background: var(--violet);
        border-radius: 10px;
    }
    ::-webkit-scrollbar-thumb:hover {
        background: var(--teal);
    }
    * {
        scrollbar-width: thin;
        scrollbar-color: var(--violet) var(--ink);
    }
    </style>
""",
    unsafe_allow_html=True,
)


def film_divider():
  """Sprocket-hole divider — the app's signature element, standing in for st.markdown('---')."""
  st.markdown(
      '<div class="film-divider">' + "<span></span>" * 22 + "</div>",
      unsafe_allow_html=True,
  )

import shutil
import subprocess
import base64

try:
  import requests
except ImportError:
  requests = None

FILE_PATH = "Blu-ray_Collection_Tracker_v5.0.xlsx"
STAR_OPTIONS = ["⭐", "⭐⭐", "⭐⭐⭐", "⭐⭐⭐⭐", "⭐⭐⭐⭐⭐"]

# BBFC (British Board of Film Classification) ratings, in age order. "12A"
# and "12" are the same age tier (cinema vs. home-media label) so they rank
# equally for filtering purposes.
BBFC_RATINGS = ["", "U", "PG", "12A", "12", "15", "18"]
BBFC_RANK = {"U": 0, "PG": 1, "12A": 2, "12": 2, "15": 3, "18": 4}



def try_recalculate(filepath):
  """Best-effort recalculation of formulas (Dashboard stats, Franchise %,
  Consensus scores) using LibreOffice, since openpyxl can't compute formulas
  itself and wipes their cached values on save. Silently does nothing if
  LibreOffice ('soffice') isn't installed -- Excel will recalculate those
  cells automatically the next time you open the file there anyway."""
  soffice = shutil.which("soffice") or shutil.which("libreoffice")
  if not soffice:
    return False
  try:
    out_dir = "recalc_tmp"
    subprocess.run(
        [soffice, "--headless", "--convert-to", "xlsx", "--outdir", out_dir, filepath],
        capture_output=True,
        timeout=30,
        check=True,
    )
    converted = f"{out_dir}/{filepath.rsplit('/', 1)[-1]}"
    import os
    if os.path.exists(converted):
      shutil.move(converted, filepath)
      shutil.rmtree(out_dir, ignore_errors=True)
      return True
  except Exception:
    pass
  return False


def github_sync_configured():
  """True if GitHub persistence secrets have been set up (only relevant when
  deployed to the cloud -- harmless no-op when running locally)."""
  try:
    return "github" in st.secrets and "token" in st.secrets["github"]
  except Exception:
    return False


def tmdb_configured():
  """True if a free TMDb API key has been added to secrets. TMDb powers
  auto-fill on Add and the plot line on Tonight's Pick -- everything using
  it degrades gracefully to 'not shown' when this is False."""
  try:
    return "tmdb" in st.secrets and "api_key" in st.secrets["tmdb"]
  except Exception:
    return False


@st.cache_data(ttl=3600, show_spinner=False)
def tmdb_lookup_cached(title, year=None):
  """Searches TMDb for a title and pulls genre/director/runtime/plot/release
  year/BBFC certificate. Cached for an hour since this data barely changes.
  Returns None on any failure or no-match -- always safe to fall back to
  manual entry."""
  if requests is None or not tmdb_configured():
    return None
  try:
    api_key = st.secrets["tmdb"]["api_key"]
    params = {"api_key": api_key, "query": title}
    if year:
      params["year"] = int(year)
    search_resp = requests.get(
        "https://api.themoviedb.org/3/search/movie", params=params, timeout=8
    )
    results = search_resp.json().get("results", [])
    if not results:
      return None

    movie_id = results[0]["id"]
    detail_resp = requests.get(
        f"https://api.themoviedb.org/3/movie/{movie_id}",
        # release_dates piggybacks on this same call -- it's where the UK
        # (BBFC) age certificate lives, no extra request needed.
        params={"api_key": api_key, "append_to_response": "credits,release_dates"},
        timeout=8,
    )
    detail = detail_resp.json()

    genres = ", ".join(g["name"] for g in detail.get("genres", []))
    crew = detail.get("credits", {}).get("crew", [])
    director = next((c["name"] for c in crew if c.get("job") == "Director"), "")
    cast = detail.get("credits", {}).get("cast", [])
    cast_sorted = sorted(cast, key=lambda c: c.get("order", 999))
    top_cast = [c["name"] for c in cast_sorted[:5] if c.get("name")]
    # Full cast (capped at 30 to avoid pathological ensemble-film sizes) --
    # only used to keep the Actors summary sheet comprehensive, not written
    # into Collection's Actor 1-5 columns, which stay top-5-only.
    full_cast = [c["name"] for c in cast_sorted[:30] if c.get("name")]
    release_date = detail.get("release_date", "") or ""
    release_year = int(release_date[:4]) if release_date[:4].isdigit() else None

    bbfc_certificate = None
    countries = detail.get("release_dates", {}).get("results", [])
    gb_entry = next((c for c in countries if c.get("iso_3166_1") == "GB"), None)
    if gb_entry:
      for d in gb_entry.get("release_dates", []):
        cert = (d.get("certification") or "").strip()
        if cert:
          bbfc_certificate = cert
          break

    return {
        "id": movie_id,
        "genre": genres,
        "director": director,
        "cast": top_cast,
        "full_cast": full_cast,
        "runtime": detail.get("runtime"),
        "poster_path": detail.get("poster_path"),
        "plot": detail.get("overview", ""),
        "year": release_year,
        "release_date": release_date,
        "certificate": bbfc_certificate,
    }
  except Exception:
    return None


@st.cache_data(ttl=86400, show_spinner=False)
def tmdb_recommendations_cached(movie_id):
  """Films similar to the given TMDb movie ID. Returns a list of
  {'title':..., 'year':..., 'poster_path':...} dicts, or [] on failure --
  used to power the 'Because you loved...' Wishlist suggestions."""
  if requests is None or not tmdb_configured():
    return []
  try:
    api_key = st.secrets["tmdb"]["api_key"]
    resp = requests.get(
        f"https://api.themoviedb.org/3/movie/{movie_id}/recommendations",
        params={"api_key": api_key},
        timeout=8,
    )
    results = resp.json().get("results", [])
    out = []
    for r in results:
      release_date = r.get("release_date", "") or ""
      out.append({
          "title": r.get("title", ""),
          "year": int(release_date[:4]) if release_date[:4].isdigit() else None,
          "poster_path": r.get("poster_path"),
      })
    return out
  except Exception:
    return []


@st.cache_data(ttl=86400, show_spinner=False)
def tmdb_trailer_cached(movie_id):
  """YouTube video key for the film's trailer, or None. Prefers an official
  Trailer, falls back to any Trailer, then a Teaser -- returns None rather
  than a mismatched clip if nothing suitable is found."""
  if requests is None or not tmdb_configured():
    return None
  try:
    api_key = st.secrets["tmdb"]["api_key"]
    resp = requests.get(
        f"https://api.themoviedb.org/3/movie/{movie_id}/videos",
        params={"api_key": api_key},
        timeout=8,
    )
    results = [v for v in resp.json().get("results", []) if v.get("site") == "YouTube"]

    official_trailer = next((v for v in results if v.get("type") == "Trailer" and v.get("official")), None)
    if official_trailer:
      return official_trailer["key"]

    any_trailer = next((v for v in results if v.get("type") == "Trailer"), None)
    if any_trailer:
      return any_trailer["key"]

    teaser = next((v for v in results if v.get("type") == "Teaser"), None)
    if teaser:
      return teaser["key"]
  except Exception:
    pass
  return None


@st.cache_data(ttl=86400, show_spinner=False)
def tmdb_release_date_cached(title, year=None, region="GB"):
  """Used by 'On This Day' -- finds the film via search, then pulls the
  actual UK release date from TMDb's per-country release_dates endpoint
  (search's own release_date field is the primary/global date, usually US,
  which is why Transformers: The Movie was showing its June 1986 US date
  instead of its December 1986 UK one). Prefers a Theatrical release entry,
  falling back to Limited Theatrical, then Premiere, then whatever's there.
  Returns {'release_date': ..., 'poster_path': ...} -- the poster comes free
  from the same search call, no extra request needed. Cached for a day since
  a whole-collection scan is comparatively expensive."""
  if requests is None or not tmdb_configured():
    return None
  try:
    api_key = st.secrets["tmdb"]["api_key"]
    params = {"api_key": api_key, "query": title}
    if year:
      params["year"] = int(year)
    resp = requests.get(
        "https://api.themoviedb.org/3/search/movie", params=params, timeout=8
    )
    results = resp.json().get("results", [])
    if not results:
      return None
    movie_id = results[0]["id"]
    fallback_date = results[0].get("release_date") or None
    poster_path = results[0].get("poster_path")

    rd_resp = requests.get(
        f"https://api.themoviedb.org/3/movie/{movie_id}/release_dates",
        params={"api_key": api_key},
        timeout=8,
    )
    countries = rd_resp.json().get("results", [])
    region_entry = next((c for c in countries if c.get("iso_3166_1") == region), None)

    release_date = fallback_date
    if region_entry:
      dates = region_entry.get("release_dates", [])
      # TMDb release types: 1=Premiere, 2=Theatrical (limited), 3=Theatrical,
      # 4=Digital, 5=Physical, 6=TV -- prefer an actual cinema release date.
      for preferred_type in (3, 2, 1):
        match = next((d for d in dates if d.get("type") == preferred_type), None)
        if match and match.get("release_date"):
          release_date = match["release_date"][:10]
          break
      else:
        dated_entries = sorted(d["release_date"] for d in dates if d.get("release_date"))
        if dated_entries:
          release_date = dated_entries[0][:10]

    return {"release_date": release_date, "poster_path": poster_path}
  except Exception:
    pass
  return None


def sync_to_github(filepath):
  """Commits the workbook back to your GitHub repo so edits survive a
  Streamlit Cloud restart or redeploy -- the container's own disk isn't
  guaranteed to persist, but your repo is. Does nothing if GitHub secrets
  aren't configured (e.g. running locally on your PC)."""
  if requests is None or not github_sync_configured():
    return False
  try:
    cfg = st.secrets["github"]
    token = cfg["token"]
    repo = cfg["repo"]  # e.g. "yourusername/blu-ray-hub"
    branch = cfg.get("branch", "main")
    path = cfg.get("path", filepath)

    api_url = f"https://api.github.com/repos/{repo}/contents/{path}"
    headers = {
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github+json",
    }

    # GitHub requires the current file's sha to update it
    get_resp = requests.get(api_url, headers=headers, params={"ref": branch}, timeout=10)
    sha = get_resp.json().get("sha") if get_resp.status_code == 200 else None

    with open(filepath, "rb") as f:
      content_b64 = base64.b64encode(f.read()).decode("utf-8")

    payload = {
        "message": f"Update collection — {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}",
        "content": content_b64,
        "branch": branch,
    }
    if sha:
      payload["sha"] = sha

    put_resp = requests.put(api_url, headers=headers, json=payload, timeout=15)
    return put_resp.status_code in (200, 201)
  except Exception:
    return False


def backup_to_github(filepath, reason="backup"):
  """Pushes a timestamped snapshot of the current workbook to a backups/
  folder in your repo, separate from the live file -- used before
  destructive operations (deleting a film, bulk backfills) so there's
  always a way back if something goes wrong. Always writes as a brand-new
  file (unique timestamped name), so unlike sync_to_github this never needs
  to look up an existing sha first. Silently does nothing if GitHub isn't
  configured -- this is a bonus safety net, not something that should ever
  block the actual operation from completing."""
  if requests is None or not github_sync_configured():
    return False
  try:
    cfg = st.secrets["github"]
    token = cfg["token"]
    repo = cfg["repo"]
    branch = cfg.get("branch", "main")

    timestamp = pd.Timestamp.now().strftime("%Y%m%d-%H%M%S")
    safe_reason = re.sub(r"[^a-zA-Z0-9_-]", "-", reason)
    backup_path = f"backups/{timestamp}-{safe_reason}.xlsx"

    api_url = f"https://api.github.com/repos/{repo}/contents/{backup_path}"
    headers = {
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github+json",
    }

    with open(filepath, "rb") as f:
      content_b64 = base64.b64encode(f.read()).decode("utf-8")

    payload = {
        "message": f"Backup before {reason} — {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}",
        "content": content_b64,
        "branch": branch,
    }

    put_resp = requests.put(api_url, headers=headers, json=payload, timeout=15)
    return put_resp.status_code in (200, 201)
  except Exception:
    return False


def save_and_sync(filepath):
  """The full save pipeline: recalculate formulas, then push to GitHub for
  durable cloud storage. Call this right after book.save(filepath)."""
  try_recalculate(filepath)
  synced = sync_to_github(filepath)
  if github_sync_configured() and not synced:
    st.warning(
        "Saved locally, but couldn't sync to GitHub — your change may not "
        "survive the app restarting. Check your GitHub token/secrets.",
        icon="⚠️",
    )


def decode_barcode(image_bytes):
  """Reads a UPC/EAN barcode from a photo. Returns the code as a string, or
  None if nothing could be decoded or the pyzbar library isn't installed."""
  try:
    from pyzbar.pyzbar import decode as zbar_decode
    from PIL import Image
    import io
    image = Image.open(io.BytesIO(image_bytes))
    results = zbar_decode(image)
    if results:
      return results[0].data.decode("utf-8")
  except ImportError:
    st.error(
        "Barcode scanning needs the pyzbar package installed "
        "(see requirements.txt / packages.txt)."
    )
  except Exception:
    pass
  return None


def clean_title_for_search(title):
  """Strips retailer-catalog junk like '(Blu-ray Disc)' or '[4K UHD Steelbook]'
  off a title before using it as a search query -- UPC product titles are
  full of this and it breaks a TMDb match otherwise."""
  cleaned = re.sub(
      r"[\(\[][^\)\]]*?(blu-?ray|dvd|4k|uhd|disc|steelbook|edition|region\s?[a-z0-9]+)[^\)\]]*?[\)\]]",
      "",
      title,
      flags=re.IGNORECASE,
  )
  return cleaned.strip(" -–—")


def find_possible_duplicates(title, threshold=0.82):
  """Checks whether a title looks like something already owned -- an exact
  case-insensitive match, plus near-matches via fuzzy string comparison to
  catch things like punctuation or minor typo differences. Returns a list
  of (existing_title, existing_year) tuples. This is advisory only (shown
  as a warning before adding), never blocks the add -- upgrades, re-buys,
  and similarly-named different films are all legitimate."""
  if "valid_collection" not in globals() or "Title" not in valid_collection.columns or not title:
    return []
  existing = valid_collection[["Title", "Year"]].dropna(subset=["Title"])
  title_norm = title.strip().lower()
  matches = []
  for _, row in existing.iterrows():
    existing_title = str(row["Title"])
    existing_norm = existing_title.strip().lower()
    if existing_norm == title_norm:
      matches.append((existing_title, row.get("Year")))
      continue
    ratio = difflib.SequenceMatcher(None, title_norm, existing_norm).ratio()
    if ratio >= threshold:
      matches.append((existing_title, row.get("Year")))
  return matches


def lookup_barcode(code):
  """Looks up a UPC/EAN code via UPCitemdb's free trial endpoint (no API key
  needed). Returns {'title': ..., 'year': int or None} on a match, or None
  if nothing was found or the request failed -- always safe to fall back to
  manual entry. UPCitemdb doesn't have a dedicated release-year field, so
  the year is a best-effort guess pulled out of the title/description text,
  and may well be missing or wrong -- worth double-checking either way."""
  if requests is None:
    return None
  try:
    resp = requests.get(
        "https://api.upcitemdb.com/prod/trial/lookup",
        params={"upc": code},
        timeout=10,
    )
    data = resp.json()
    items = data.get("items", [])
    if not items:
      return None

    item = items[0]
    title = item.get("title", "")
    description = item.get("description", "") or ""

    year = None
    year_match = re.search(r"\b(19[5-9]\d|20[0-4]\d)\b", f"{title} {description}")
    if year_match:
      year = int(year_match.group(1))

    return {"title": title, "year": year}
  except Exception:
    pass
  return None


# ----------------------------- helpers -----------------------------------

def safe_year(value):
  """Turn a possibly-float/possibly-string year into a clean display string."""
  if pd.isna(value):
    return ""
  try:
    return str(int(float(value)))
  except (ValueError, TypeError):
    return str(value)


def star_index_from_rating(rating_val, default=2):
  """Map a stored rating (either '⭐⭐⭐' or a number) to a 0-4 index."""
  if pd.isna(rating_val):
    return default
  rating_str = str(rating_val).strip()
  if "⭐" in rating_str:
    return max(0, min(4, rating_str.count("⭐") - 1))
  try:
    return max(0, min(4, int(float(rating_val)) - 1))
  except (ValueError, TypeError):
    return default


def is_watched(value):
  return str(value).strip().lower() in ["yes", "y", "true"]


def watched_display(value):
  """Turn a raw Watched cell (which may be NaN/blank) into 'Yes' or 'No'."""
  return "Yes" if is_watched(value) else "No"


def format_badge_html(format_value):
  """Small colored pill for Format — violet for Blu-ray, teal for 4K UHD,
  matching the app's laser-blue / disc-shimmer palette."""
  if pd.isna(format_value):
    return ""
  label = str(format_value).strip()
  lower = label.lower()
  if "4k" in lower:
    css_class = "uhd4k"
  elif "blu" in lower:
    css_class = "bluray"
  else:
    css_class = "other"
  return f'<span class="format-badge {css_class}">{label}</span>'


def retailer_search_links(title):
  """Search-page links for a title, one tap instead of typing it out each
  time. HMV/Zavvi don't have a documented simple search URL, so those go via
  a site-scoped Google search (same idea as the IMDb/RT lookup links already
  in your Ratings sheet); Amazon and CEX both have confirmed, documented
  search URL formats so those link directly."""
  quoted = urllib.parse.quote_plus(title)
  return {
      "HMV": f"https://www.google.com/search?q=site:hmv.com+{quoted}",
      "Zavvi": f"https://www.google.com/search?q=site:zavvi.com+{quoted}",
      "Amazon": f"https://www.amazon.co.uk/s?k={quoted}",
      "CEX": f"https://uk.webuy.com/search?stext={quoted}",
  }


def rating_stars_display(rating_val):
  """Turn a raw Rating cell into a star string, or a neutral placeholder if unset."""
  if pd.isna(rating_val) or str(rating_val).strip() == "":
    return "Not rated yet"
  rating_str = str(rating_val).strip()
  if "⭐" in rating_str:
    return rating_str
  try:
    return STAR_OPTIONS[max(0, min(4, int(float(rating_val)) - 1))]
  except (ValueError, TypeError):
    return "Not rated yet"


def curated_browse_view(df):
  """A phone-friendly slice of the collection: just the columns someone
  browsing would actually want to see, with clean values -- not the full
  26-column spreadsheet. Keeps the sheet's own row order (shelf/Film ID
  order) rather than re-sorting, per preference."""
  wanted_cols = ["Title", "Year", "Format", "Watched", "Rating (1-5)", "Date Watched"]
  cols_present = [c for c in wanted_cols if c in df.columns]
  view = df[cols_present].copy()

  if "Year" in view.columns:
    view["Year"] = view["Year"].apply(safe_year)
  if "Watched" in view.columns:
    view["Watched"] = view["Watched"].apply(watched_display)
  if "Rating (1-5)" in view.columns:
    view["Rating (1-5)"] = view["Rating (1-5)"].apply(rating_stars_display)
    view = view.rename(columns={"Rating (1-5)": "Rating"})
  if "Date Watched" in view.columns:
    view["Date Watched"] = view["Date Watched"].apply(format_date_watched)
    view = view.rename(columns={"Date Watched": "Last Watched"})

  return view.reset_index(drop=True)


def style_format_column(df):
  """Colors the Format column violet for Blu-ray / teal for 4K UHD, matching
  the app's palette. Returns a Styler; falls back to the plain DataFrame if
  the installed pandas version doesn't support cell styling."""
  if "Format" not in df.columns:
    return df

  def _color(value):
    lower = str(value).lower()
    if "4k" in lower:
      return "color: #2DD4BF; font-weight: 700;"
    if "blu" in lower:
      return "color: #A79CFF; font-weight: 700;"
    return ""

  try:
    return df.style.map(_color, subset=["Format"])
  except AttributeError:
    try:
      return df.style.applymap(_color, subset=["Format"])
    except Exception:
      return df


def find_row_by_id(sheet, id_column, target_id, header_row=4):
  """Scan a sheet for the row whose id_column matches target_id."""
  for r in range(header_row, sheet.max_row + 1):
    if str(sheet.cell(row=r, column=id_column).value).strip() == str(
        target_id
    ).strip():
      return r
  return None


def get_col_index(sheet, header_name, header_row=4):
  """Finds a column's position by its header text instead of a hardcoded
  number, so writes stay correct even if the sheet's columns get reordered."""
  for c in range(1, sheet.max_column + 1):
    if str(sheet.cell(row=header_row, column=c).value).strip() == header_name:
      return c
  return None


def ensure_column(sheet, header_name, header_row=4):
  """Like get_col_index, but creates the column (writing the header text)
  if it doesn't exist yet -- used for 'Date Watched', which isn't part of
  the original spreadsheet."""
  col = get_col_index(sheet, header_name, header_row)
  if col:
    return col
  new_col = sheet.max_column + 1
  sheet.cell(row=header_row, column=new_col, value=header_name)
  return new_col


def format_date_watched(value):
  """Turns a stored Date Watched value into a clean dd/mm/yyyy string, or
  '' if it's blank/unset."""
  if pd.isna(value) or str(value).strip() == "":
    return ""
  if isinstance(value, (datetime.date, datetime.datetime)):
    return value.strftime("%d/%m/%Y")
  try:
    return pd.to_datetime(value).strftime("%d/%m/%Y")
  except Exception:
    return str(value)


def cast_fields_from_string(cast_text):
  """Splits a comma-joined cast string into the Actor 1..Actor 5 columns
  the sheet actually uses. Extra names beyond 5 are dropped."""
  names = [n.strip() for n in cast_text.split(",") if n.strip()][:5]
  return {f"Actor {i + 1}": name for i, name in enumerate(names)}


def update_person_sheet(book, sheet_name, name_col_header, person_name, film_title, film_year, film_id):
  """Keeps the Directors/Actors summary sheets in sync when a film is added.
  These are static pre-computed values baked into the workbook, not
  formulas, so nothing recalculates them automatically -- confirmed by
  inspecting the raw cells rather than assumed. Updates the matching
  person's row if one exists, or adds a new one."""
  if sheet_name not in book.sheetnames or not person_name:
    return
  sheet = book[sheet_name]
  header_row = 4

  name_col = get_col_index(sheet, name_col_header, header_row)
  owned_col = get_col_index(sheet, "Films Owned", header_row)
  earliest_col = get_col_index(sheet, "Earliest Film", header_row)
  latest_col = get_col_index(sheet, "Latest Film", header_row)
  films_col = get_col_index(sheet, "Films in Collection", header_row)
  ids_col = get_col_index(sheet, "Film IDs", header_row)
  if not name_col:
    return

  target_row = None
  for r in range(header_row + 1, sheet.max_row + 1):
    cell_val = sheet.cell(row=r, column=name_col).value
    if cell_val and str(cell_val).strip().lower() == person_name.strip().lower():
      target_row = r
      break

  try:
    year_int = int(float(film_year)) if film_year else None
  except (TypeError, ValueError):
    year_int = None

  if target_row:
    current_films_val = sheet.cell(row=target_row, column=films_col).value or "" if films_col else ""
    existing_titles = [t.strip() for t in str(current_films_val).split(",") if t.strip()]
    already_has_film = bool(film_title) and film_title in existing_titles

    # Only increment/append when this is genuinely a new film for this
    # person -- otherwise re-running this (e.g. a bulk backfill across
    # films that are already correctly counted) would double-count them.
    if not already_has_film:
      if owned_col:
        current = sheet.cell(row=target_row, column=owned_col).value or 0
        sheet.cell(row=target_row, column=owned_col, value=int(current) + 1)
      if films_col and film_title:
        existing_titles.append(film_title)
        sheet.cell(row=target_row, column=films_col, value=", ".join(existing_titles))
      if ids_col:
        current = sheet.cell(row=target_row, column=ids_col).value or ""
        existing_ids = [i.strip() for i in str(current).split(",") if i.strip()]
        if film_id and film_id not in existing_ids:
          existing_ids.append(film_id)
          sheet.cell(row=target_row, column=ids_col, value=", ".join(existing_ids))

    # Earliest/Latest are safe to recompute regardless -- min/max is
    # naturally idempotent.
    if earliest_col and year_int:
      current = sheet.cell(row=target_row, column=earliest_col).value
      if not current or year_int < int(current):
        sheet.cell(row=target_row, column=earliest_col, value=year_int)
    if latest_col and year_int:
      current = sheet.cell(row=target_row, column=latest_col).value
      if not current or year_int > int(current):
        sheet.cell(row=target_row, column=latest_col, value=year_int)
  else:
    new_row = sheet.max_row + 1
    sheet.cell(row=new_row, column=name_col, value=person_name)
    if owned_col:
      sheet.cell(row=new_row, column=owned_col, value=1)
    if earliest_col and year_int:
      sheet.cell(row=new_row, column=earliest_col, value=year_int)
    if latest_col and year_int:
      sheet.cell(row=new_row, column=latest_col, value=year_int)
    if films_col and film_title:
      sheet.cell(row=new_row, column=films_col, value=film_title)
    if ids_col and film_id:
      sheet.cell(row=new_row, column=ids_col, value=film_id)


def backfill_missing_details(films_to_process):
  """One-time bulk enrichment for films added before TMDb auto-fill existed
  (or added manually without it). Looks each film up on TMDb in parallel,
  then fills in Genre/Director/Cast/Runtime/Notes -- but ONLY on cells that
  are genuinely blank, never overwriting anything already filled in. Also
  syncs the Directors/Actors summary sheets the same way a normal Add does.
  films_to_process is a list of (film_id, title, year) tuples. Returns
  (updated_count, no_match_count). Does one single save at the end rather
  than one per film, since this can touch hundreds of rows at once."""

  def _lookup(item):
    film_id, title, year = item
    return film_id, title, year, tmdb_lookup_cached(title, year)

  with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
    lookups = list(executor.map(_lookup, films_to_process))

  backup_to_github(FILE_PATH, reason="bulk-backfill")

  book = openpyxl.load_workbook(FILE_PATH)
  sheet = book["Collection"]

  updated = 0
  no_match = 0

  genre_col = get_col_index(sheet, "Genre")
  director_col = get_col_index(sheet, "Director")
  runtime_col = get_col_index(sheet, "Runtime (min)")
  notes_col = get_col_index(sheet, "Notes")
  cert_col = ensure_column(sheet, "BBFC Rating")
  actor_cols = [get_col_index(sheet, f"Actor {i}") for i in range(1, 6)]

  for film_id, title, year, result in lookups:
    if not result:
      no_match += 1
      continue

    target_row = find_row_by_id(sheet, id_column=1, target_id=film_id)
    if not target_row:
      no_match += 1
      continue

    changed = False

    def _fill(col, value):
      nonlocal changed
      if col and value and not sheet.cell(row=target_row, column=col).value:
        sheet.cell(row=target_row, column=col, value=value)
        changed = True

    _fill(genre_col, result.get("genre"))
    _fill(director_col, result.get("director"))
    _fill(runtime_col, result.get("runtime"))
    _fill(notes_col, result.get("plot"))
    _fill(cert_col, result.get("certificate"))

    top_cast = result.get("cast", [])
    for i, col in enumerate(actor_cols):
      if i < len(top_cast):
        _fill(col, top_cast[i])

    if result.get("director"):
      update_person_sheet(book, "Directors", "Director", result["director"], title, year, film_id)
    for actor_name in result.get("full_cast", []):
      update_person_sheet(book, "Actors", "Actor", actor_name, title, year, film_id)

    if changed:
      updated += 1

  book.save(FILE_PATH)
  save_and_sync(FILE_PATH)
  return updated, no_match


def add_to_collection(fields, full_cast=None):
  """Adds a new film to the Collection sheet. fields is a dict of
  {header_name: value} -- only columns that exist in the sheet get written,
  so this stays safe even if the sheet's structure changes. Also syncs the
  Directors/Actors summary sheets so 'Browse by Actor/Director' picks up
  the new film immediately instead of only showing stale, pre-built data.

  full_cast (optional) is the FULL cast list from TMDb, used only for the
  Actors-sheet sync -- Collection's own Actor 1-5 columns stay top-5-only
  (that's a fixed structural limit of the sheet), but the Actors summary
  sheet has no such cap, so someone like a supporting/non-top-5-billed actor
  still shows up under 'Browse by Actor' instead of being invisible. Falls
  back to just the 5 names written to Collection if not provided."""
  book = openpyxl.load_workbook(FILE_PATH)
  sheet = book["Collection"]
  next_row = sheet.max_row + 1

  film_id_col = get_col_index(sheet, "Film ID") or 1
  new_film_id = next_film_id(collection_df)
  sheet.cell(row=next_row, column=film_id_col, value=new_film_id)

  for header_name, value in fields.items():
    if value in (None, ""):
      continue
    col = ensure_column(sheet, header_name)
    if col:
      sheet.cell(row=next_row, column=col, value=value)

  film_title = fields.get("Title", "")
  film_year = fields.get("Year")
  director_name = fields.get("Director")
  if director_name:
    update_person_sheet(book, "Directors", "Director", director_name, film_title, film_year, new_film_id)

  actor_names = full_cast if full_cast else [fields.get(f"Actor {i}") for i in range(1, 6)]
  for actor_name in actor_names:
    if actor_name:
      update_person_sheet(book, "Actors", "Actor", actor_name, film_title, film_year, new_film_id)

  book.save(FILE_PATH)
  save_and_sync(FILE_PATH)


def delete_from_collection(film_id):
  """Removes a film from the Collection sheet entirely. Takes a timestamped
  GitHub backup first, since this is destructive and can't be undone
  through the app itself."""
  backup_to_github(FILE_PATH, reason=f"delete-{film_id}")

  book = openpyxl.load_workbook(FILE_PATH)
  sheet = book["Collection"]
  target_row = find_row_by_id(sheet, id_column=1, target_id=film_id)
  if not target_row:
    return False
  sheet.delete_rows(target_row)
  book.save(FILE_PATH)
  save_and_sync(FILE_PATH)
  return True


def stars_to_number(star_string):
  """Turn '⭐⭐⭐' into 3. Falls back to None if it can't be parsed."""
  count = str(star_string).count("⭐")
  return count if count > 0 else None


def append_watch_log(book, film_id, title, date_str, rating_value=None):
  """Adds one entry to the Watch Log sheet, creating it (with a header row
  matching the rest of the workbook's row-4 convention) if it doesn't exist
  yet. Unlike Collection's single 'Date Watched' cell, this keeps every
  watch as its own row, so the same film can be watched multiple times."""
  if "Watch Log" not in book.sheetnames:
    ws = book.create_sheet("Watch Log")
    ws.cell(row=4, column=1, value="Film ID")
    ws.cell(row=4, column=2, value="Title")
    ws.cell(row=4, column=3, value="Date Watched")
    ws.cell(row=4, column=4, value="Rating at Time")
  else:
    ws = book["Watch Log"]

  next_row = ws.max_row + 1
  ws.cell(row=next_row, column=1, value=film_id)
  ws.cell(row=next_row, column=2, value=title)
  ws.cell(row=next_row, column=3, value=date_str)
  numeric_rating = stars_to_number(rating_value) if rating_value else None
  if numeric_rating is not None:
    ws.cell(row=next_row, column=4, value=numeric_rating)


def save_collection_update(film_id, watched_value, rating_value, date_watched=None):
  """Writes Watched + Rating to the Collection sheet, and mirrors the
  numeric rating into the Ratings sheet's 'Your Rating /5' column so the
  two stay in sync. date_watched is only written when the film is marked
  Watched -- pass None to leave it untouched (e.g. when unmarking).

  Also logs a Watch Log entry, but ONLY when the date actually changed from
  what was already stored -- otherwise just resaving a rating edit (without
  a genuinely new watch) would create a phantom log entry every time."""
  book = openpyxl.load_workbook(FILE_PATH)
  coll_sheet = book["Collection"]
  target_row = find_row_by_id(coll_sheet, id_column=1, target_id=film_id)
  if not target_row:
    return False

  title_col = get_col_index(coll_sheet, "Title")
  film_title = coll_sheet.cell(row=target_row, column=title_col).value if title_col else ""

  coll_sheet.cell(row=target_row, column=7, value=watched_value)
  coll_sheet.cell(row=target_row, column=8, value=rating_value)

  if is_watched(watched_value) and date_watched:
    date_col = ensure_column(coll_sheet, "Date Watched")
    date_str = date_watched.strftime("%Y-%m-%d") if hasattr(date_watched, "strftime") else str(date_watched)
    old_date_val = coll_sheet.cell(row=target_row, column=date_col).value
    old_date_str = str(old_date_val)[:10] if old_date_val else None

    if date_str != old_date_str:
      append_watch_log(book, film_id, film_title, date_str, rating_value)

    coll_sheet.cell(row=target_row, column=date_col, value=date_str)

  if "Ratings" in book.sheetnames:
    ratings_sheet = book["Ratings"]
    ratings_row = find_row_by_id(ratings_sheet, id_column=1, target_id=film_id)
    numeric_rating = stars_to_number(rating_value)
    if ratings_row and numeric_rating is not None:
      ratings_sheet.cell(row=ratings_row, column=4, value=numeric_rating)

  book.save(FILE_PATH)
  save_and_sync(FILE_PATH)
  return True


# Load data with caching
@st.cache_data
def load_data():
  collection_df = pd.read_excel(FILE_PATH, sheet_name="Collection", skiprows=3)
  wishlist_df = pd.read_excel(FILE_PATH, sheet_name="Wishlist", skiprows=3)
  franchise_df = pd.read_excel(
      FILE_PATH, sheet_name="Franchise Tracker", skiprows=3
  )
  actors_df = pd.read_excel(FILE_PATH, sheet_name="Actors", skiprows=3)
  directors_df = pd.read_excel(FILE_PATH, sheet_name="Directors", skiprows=3)
  awards_df = pd.read_excel(FILE_PATH, sheet_name="Awards", skiprows=3)
  ratings_df = pd.read_excel(FILE_PATH, sheet_name="Ratings", skiprows=3)

  # Watch Log doesn't exist in the original spreadsheet -- it's created the
  # first time a watch gets logged, so read it defensively.
  wb_check = openpyxl.load_workbook(FILE_PATH, read_only=True)
  has_watch_log = "Watch Log" in wb_check.sheetnames
  wb_check.close()
  if has_watch_log:
    watch_log_df = pd.read_excel(FILE_PATH, sheet_name="Watch Log", skiprows=3)
  else:
    watch_log_df = pd.DataFrame(columns=["Film ID", "Title", "Date Watched", "Rating at Time"])

  return (
      collection_df,
      wishlist_df,
      franchise_df,
      actors_df,
      directors_df,
      awards_df,
      ratings_df,
      watch_log_df,
  )


def split_multi_value_counts(series):
  """Split comma-separated cells (e.g. 'Action, Comedy') and count each value."""
  exploded = (
      series.dropna()
      .astype(str)
      .str.split(",")
      .explode()
      .str.strip()
  )
  exploded = exploded[exploded != ""]
  return exploded.value_counts()


def next_film_id(collection_df):
  """Generate the next Film ID in the F0001-style sequence."""
  existing = (
      collection_df["Film ID"]
      .dropna()
      .astype(str)
      .str.extract(r"F(\d+)", expand=False)
      .dropna()
      .astype(int)
  )
  next_num = (existing.max() + 1) if not existing.empty else 1
  return f"F{next_num:04d}"


def add_to_wishlist(title, priority=3, target_price=None, notes=""):
  """Adds a film to the Wishlist sheet. Shared by the Wishlist form and the
  'Add to Wishlist' shortcut on a zero-result Search."""
  book = openpyxl.load_workbook(FILE_PATH)
  sheet = book["Wishlist"]
  next_row = sheet.max_row + 1
  sheet.cell(row=next_row, column=1, value=title)
  sheet.cell(row=next_row, column=3, value=priority)
  if target_price is not None:
    sheet.cell(row=next_row, column=4, value=target_price)
  sheet.cell(row=next_row, column=8, value="No")
  sheet.cell(row=next_row, column=9, value=notes)
  book.save(FILE_PATH)
  save_and_sync(FILE_PATH)


try:
  (
      collection_df,
      wishlist_df,
      franchise_df,
      actors_df,
      directors_df,
      awards_df,
      ratings_df,
      watch_log_df,
  ) = load_data()

  valid_collection = collection_df.dropna(subset=["Title"]).copy()
  valid_wishlist = wishlist_df.dropna(subset=["Title"]).copy()
  total_collection = len(valid_collection)
  total_wishlist = len(valid_wishlist)

  # Detect an optional price/cost column so the value stat only shows if it exists
  price_col = next(
      (c for c in ["Price", "Price (£)", "Cost", "Cost (£)"]
       if c in valid_collection.columns),
      None,
  )

  # App Header
  st.markdown(
      "<div class='app-title'>🎬 Conor's Blu-ray Hub</div>",
      unsafe_allow_html=True,
  )
  st.markdown(
      "<div class='app-subtitle'>Mobile Media Companion</div>",
      unsafe_allow_html=True,
  )

  # Injects a proper home-screen icon for iOS/Android "Add to Home Screen",
  # since Streamlit's page_icon only sets the browser tab favicon, not the
  # icon used when saving the app to a phone's home screen. Works by
  # reaching into the parent page's <head> from an invisible iframe -- a
  # known Streamlit community technique, not an official API, so it's
  # best-effort rather than guaranteed on every browser/version.
  _icon_svg = (
      '<svg xmlns="http://www.w3.org/2000/svg" width="180" height="180" viewBox="0 0 180 180">'
      '<defs><linearGradient id="g" x1="0%" y1="0%" x2="100%" y2="100%">'
      '<stop offset="0%" stop-color="#6C5CE8"/><stop offset="100%" stop-color="#2DD4BF"/>'
      '</linearGradient></defs>'
      '<rect width="180" height="180" rx="38" fill="url(#g)"/>'
      '<circle cx="90" cy="90" r="54" fill="#0A0B14"/>'
      '<circle cx="90" cy="90" r="54" fill="none" stroke="#F0A83B" stroke-width="5"/>'
      '<circle cx="90" cy="90" r="13" fill="#F0A83B"/>'
      '</svg>'
  )
  _icon_data_uri = "data:image/svg+xml;base64," + base64.b64encode(_icon_svg.encode()).decode()
  components.html(
      f"""
      <script>
      (function() {{
        var head = window.parent.document.head;
        if (!head.querySelector("link[rel='apple-touch-icon']")) {{
          var link1 = document.createElement('link');
          link1.rel = 'apple-touch-icon';
          link1.href = '{_icon_data_uri}';
          head.appendChild(link1);
          var link2 = document.createElement('link');
          link2.rel = 'icon';
          link2.href = '{_icon_data_uri}';
          head.appendChild(link2);
        }}
      }})();
      </script>
      """,
      height=0,
  )

  if github_sync_configured():
    st.markdown(
        "<p style='text-align:center; font-size:0.75em; color:var(--teal); "
        "margin-top:4px;'>☁️ Cloud sync active</p>",
        unsafe_allow_html=True,
    )
  else:
    st.markdown(
        "<p style='text-align:center; font-size:0.75em; opacity:0.5; "
        "margin-top:4px;'>💾 Local storage only — edits may not survive a restart</p>",
        unsafe_allow_html=True,
    )
  film_divider()

  # Navigation Tabs
  app_mode = st.tabs([
      "🏠 Home",
      "🔍 Search",
      "📚 Collection",
      "🛒 Wishlist",
      "📊 Stats",
      "🏆 Extras",
      "📅 On This Day",
  ])

  # --- TAB 1: HOME & RANDOM PICKER ---
  with app_mode[0]:
    st.markdown("### 📊 Library Stats")

    watched_count = valid_collection["Watched"].apply(is_watched).sum() if "Watched" in valid_collection.columns else 0
    unwatched_count = total_collection - watched_count

    stat_cols_row1 = st.columns(2)
    with stat_cols_row1[0]:
      if st.button(f"🎬  **{total_collection}**  \nOwned", key="stat_owned", use_container_width=True):
        st.session_state["home_stat_view"] = "owned"
    with stat_cols_row1[1]:
      if st.button(f"🛒  **{total_wishlist}**  \nWishlist", key="stat_wishlist", use_container_width=True):
        st.session_state["home_stat_view"] = "wishlist"

    stat_cols_row2 = st.columns(2)
    with stat_cols_row2[0]:
      if st.button(f"✅  **{watched_count}**  \nWatched", key="stat_watched", use_container_width=True):
        st.session_state["home_stat_view"] = "watched"
    with stat_cols_row2[1]:
      if st.button(f"⬜  **{unwatched_count}**  \nUnwatched", key="stat_unwatched", use_container_width=True):
        st.session_state["home_stat_view"] = "unwatched"

    if price_col:
      total_value = pd.to_numeric(
          valid_collection[price_col], errors="coerce"
      ).sum()
      st.markdown(
          f"<p style='text-align: center; color: gray; margin-top: 4px;'>"
          f"💰 Estimated collection value: <b>£{total_value:,.2f}</b></p>",
          unsafe_allow_html=True,
      )

    active_stat_view = st.session_state.get("home_stat_view")
    if active_stat_view:
      view_titles = {
          "owned": f"🎬 All {total_collection} owned films",
          "wishlist": f"🛒 All {total_wishlist} wishlist films",
          "watched": f"✅ {watched_count} watched films",
          "unwatched": f"⬜ {unwatched_count} unwatched films",
      }
      hcol1, hcol2 = st.columns([4, 1])
      with hcol1:
        st.markdown(f"**{view_titles[active_stat_view]}**")
      with hcol2:
        if st.button("✕ Close", key="close_stat_view"):
          st.session_state.pop("home_stat_view", None)
          st.rerun()

      if active_stat_view == "owned":
        st.dataframe(style_format_column(curated_browse_view(valid_collection)), use_container_width=True, hide_index=True)
      elif active_stat_view == "wishlist":
        wishlist_cols = [c for c in ["Title", "Priority (1-5)", "Target Price (£)"] if c in valid_wishlist.columns]
        st.dataframe(valid_wishlist[wishlist_cols], use_container_width=True, hide_index=True)
      elif active_stat_view == "watched" and "Watched" in valid_collection.columns:
        watched_films = valid_collection[valid_collection["Watched"].apply(is_watched)]
        st.dataframe(style_format_column(curated_browse_view(watched_films)), use_container_width=True, hide_index=True)
      elif active_stat_view == "unwatched" and "Watched" in valid_collection.columns:
        unwatched_films = valid_collection[~valid_collection["Watched"].apply(is_watched)]
        st.dataframe(style_format_column(curated_browse_view(unwatched_films)), use_container_width=True, hide_index=True)

    film_divider()
    st.markdown("### 🎲 Movie Night Decider")

    rcol1, rcol2 = st.columns(2)
    with rcol1:
      formats = (
          ["All"]
          + sorted(valid_collection["Format"].dropna().unique().tolist())
          if "Format" in valid_collection.columns
          else ["All"]
      )
      selected_format_filter = st.selectbox(
          "Format", formats, key="rand_format"
      )

    with rcol2:
      selected_watched_filter = st.selectbox(
          "Status", ["All", "Unwatched Only", "Watched Only"]
      )

    selected_age_filter = st.selectbox(
        "Age rating",
        ["All", "Kids only (U/PG)", "12A and under"],
        key="rand_age_filter",
    )

    genre_options = (
        split_multi_value_counts(valid_collection["Genre"]).index.tolist()
        if "Genre" in valid_collection.columns else []
    )
    selected_genres_filter = st.multiselect("Genre", genre_options, key="rand_genre_filter")

    double_feature = st.toggle("🎬 Double Feature (pick 2 films)", value=False)

    pool_df = valid_collection.copy()
    if selected_format_filter != "All" and "Format" in pool_df.columns:
      pool_df = pool_df[pool_df["Format"] == selected_format_filter]

    if "Watched" in pool_df.columns:
      if selected_watched_filter == "Unwatched Only":
        pool_df = pool_df[~pool_df["Watched"].apply(is_watched)]
      elif selected_watched_filter == "Watched Only":
        pool_df = pool_df[pool_df["Watched"].apply(is_watched)]

    if selected_age_filter != "All" and "BBFC Rating" in pool_df.columns:
      max_rank = 1 if selected_age_filter == "Kids only (U/PG)" else 2
      age_ranks = pool_df["BBFC Rating"].map(BBFC_RANK)
      pool_df = pool_df[age_ranks.notna() & (age_ranks <= max_rank)]

    if selected_genres_filter and "Genre" in pool_df.columns:
      def _genre_match(cell):
        cell_genres = [g.strip() for g in str(cell).split(",")]
        return any(g in cell_genres for g in selected_genres_filter)
      pool_df = pool_df[pool_df["Genre"].apply(_genre_match)]

    st.markdown(
        f"<p style='text-align: center; font-size: 0.9em; color:"
        f" gray;'>Available pool: <b>{len(pool_df)}</b> films</p>",
        unsafe_allow_html=True,
    )

    pick_label = "🎰 Pick a Double Feature!" if double_feature else "🎰 Pick a Random Movie!"
    if st.button(pick_label, type="primary"):
      needed = 2 if double_feature else 1
      if len(pool_df) >= needed:
        picked = pool_df.sample(n=needed)
        st.session_state["picked_movie_ids"] = picked.get(
            "Film ID", picked.index.to_series()
        ).astype(str).tolist()
      else:
        st.warning("Not enough movies match your filters.")
        st.session_state.pop("picked_movie_ids", None)

    if "picked_movie_ids" in st.session_state and "Film ID" in valid_collection.columns:
      picked_rows = valid_collection[
          valid_collection["Film ID"].astype(str).isin(
              st.session_state["picked_movie_ids"]
          )
      ]

      for _, picked_movie in picked_rows.iterrows():
        p_title = picked_movie.get("Title", "Unknown")
        p_year_str = safe_year(picked_movie.get("Year", ""))
        p_format_badge = format_badge_html(picked_movie.get("Format", "Blu-ray"))
        p_genre = picked_movie.get("Genre", "N/A")
        p_director = picked_movie.get("Director", "N/A")
        p_id = picked_movie.get("Film ID", "")
        p_watched_raw = picked_movie.get("Watched", "No")
        p_watched = watched_display(p_watched_raw)
        p_rating_display = rating_stars_display(picked_movie.get("Rating (1-5)"))
        p_date_watched = format_date_watched(picked_movie.get("Date Watched")) if "Date Watched" in picked_movie.index else ""
        p_last_watched_html = f"<br><b>Last watched:</b> {p_date_watched}" if p_date_watched else ""

        p_plot_html = ""
        p_trailer_key = None
        if tmdb_configured():
          tmdb_pick_data = tmdb_lookup_cached(p_title, picked_movie.get("Year"))
          if tmdb_pick_data and tmdb_pick_data.get("plot"):
            p_plot_html = (
                f'<p style="font-size: 0.85em; margin-top: 10px; opacity: 0.75; '
                f'font-style: italic;">{tmdb_pick_data["plot"]}</p>'
            )
          if tmdb_pick_data and tmdb_pick_data.get("id"):
            p_trailer_key = tmdb_trailer_cached(tmdb_pick_data["id"])

        st.markdown(
            f"""
                <div class="winner-box">
                    <h3 style="color: #F0A83B; margin-bottom: 2px;">🎉 Tonight's Pick:</h3>
                    <h2 style="margin: 0px;">{p_title} ({p_year_str})</h2>
                    <p style="font-size: 0.95em; margin-top: 8px; opacity: 0.85;">
                        {p_format_badge} &nbsp;<b>Director:</b> {p_director}<br>
                        <b>Genre:</b> {p_genre} | <b>Watched:</b> {p_watched}<br>
                        <b>Rating:</b> <span class="stars-display">{p_rating_display}</span>{p_last_watched_html}
                    </p>
                    {p_plot_html}
                </div>
                """,
            unsafe_allow_html=True,
        )

        if p_trailer_key:
          st.link_button("▶️ Watch Trailer", f"https://www.youtube.com/watch?v={p_trailer_key}")

        with st.form(f"quick_picker_update_form_{p_id}"):
          st.markdown(f"**Quick Update — {p_title}:**")
          uq_col1, uq_col2 = st.columns(2)
          with uq_col1:
            quick_watched = st.selectbox(
                "Watched",
                ["Yes", "No"],
                index=0 if is_watched(p_watched) else 1,
                key=f"qp_watched_{p_id}",
            )
          with uq_col2:
            quick_stars = st.selectbox(
                "Rating",
                STAR_OPTIONS,
                index=2,
                key=f"qp_stars_{p_id}",
            )
          quick_date_watched = st.date_input(
              "Date watched (optional)",
              value=pd.to_datetime(picked_movie.get("Date Watched")).date()
                    if "Date Watched" in picked_movie.index and pd.notna(picked_movie.get("Date Watched"))
                    else datetime.date.today(),
              key=f"qp_date_{p_id}",
          )

          if st.form_submit_button("💾 Save to Excel"):
            if save_collection_update(p_id, quick_watched, quick_stars, quick_date_watched):
              st.cache_data.clear()
              st.success(f"Updated '{p_title}' successfully!")
              if quick_stars == STAR_OPTIONS[-1]:
                st.balloons()
              st.rerun()
            else:
              st.error("Couldn't find that film in the sheet to update.")

  # --- TAB 2: SEARCH ---
  with app_mode[1]:
    st.subheader("🔍 Search Library")
    scol1, scol2 = st.columns([2, 1])
    with scol1:
      query = st.text_input("Keyword search:")
    with scol2:
      search_column = st.selectbox(
          "In column", ["All"] + collection_df.columns.tolist()
      )

    with st.expander("🔎 Filters"):
      fcol1, fcol2 = st.columns(2)
      with fcol1:
        genre_options = (
            split_multi_value_counts(valid_collection["Genre"]).index.tolist()
            if "Genre" in valid_collection.columns else []
        )
        selected_genres = st.multiselect("Genre", genre_options)
      with fcol2:
        if "Year" in valid_collection.columns:
          decades_available = sorted(
              pd.to_numeric(valid_collection["Year"], errors="coerce")
              .dropna().apply(lambda y: int(y) // 10 * 10).unique().tolist()
          )
          decade_options = ["All"] + [f"{d}s" for d in decades_available]
        else:
          decade_options = ["All"]
        selected_decade = st.selectbox("Decade", decade_options)

      fcol3, fcol4 = st.columns(2)
      with fcol3:
        format_options = (
            ["All"] + sorted(valid_collection["Format"].dropna().unique().tolist())
            if "Format" in valid_collection.columns else ["All"]
        )
        selected_format_filter2 = st.selectbox("Format", format_options, key="search_filter_format")
      with fcol4:
        selected_watched_filter2 = st.selectbox("Status", ["All", "Watched", "Unwatched"], key="search_filter_watched")

    filters_active = bool(selected_genres) or selected_decade != "All" or selected_format_filter2 != "All" or selected_watched_filter2 != "All"

    if query or filters_active:
      results = collection_df.dropna(subset=["Title"]).copy()

      if query:
        if search_column == "All":
          mask = results.astype(str).apply(
              lambda x: x.str.contains(query, case=False, na=False)
          ).any(axis=1)
        else:
          mask = (
              results[search_column]
              .astype(str)
              .str.contains(query, case=False, na=False)
          )
        results = results[mask]

      if selected_genres and "Genre" in results.columns:
        def _genre_match(cell):
          cell_genres = [g.strip() for g in str(cell).split(",")]
          return any(g in cell_genres for g in selected_genres)
        results = results[results["Genre"].apply(_genre_match)]

      if selected_decade != "All" and "Year" in results.columns:
        decade_start = int(selected_decade[:-1])
        years = pd.to_numeric(results["Year"], errors="coerce")
        results = results[(years >= decade_start) & (years < decade_start + 10)]

      if selected_format_filter2 != "All" and "Format" in results.columns:
        results = results[results["Format"] == selected_format_filter2]

      if selected_watched_filter2 != "All" and "Watched" in results.columns:
        if selected_watched_filter2 == "Watched":
          results = results[results["Watched"].apply(is_watched)]
        else:
          results = results[~results["Watched"].apply(is_watched)]

      if len(results) == 0:
        no_match_msg = f"No matches for '{query}'" if query else "No matches for these filters"
        st.warning(f"{no_match_msg} in your collection.")
        if query and st.button(f"🛒 Add '{query}' to Wishlist", key="search_add_to_wishlist"):
          add_to_wishlist(query)
          st.cache_data.clear()
          st.success(f"Added '{query}' to your Wishlist!")
          st.rerun()
      else:
        st.success(f"Found {len(results)} matches:")
        st.dataframe(style_format_column(curated_browse_view(results)), use_container_width=True, hide_index=True)
    else:
      st.info("Type a keyword or set a filter above to find a film.")

  # --- TAB 3: COLLECTION & UPDATE ---
  with app_mode[2]:
    st.subheader("📚 Update Status & Rating")

    valid_collection_sorted = valid_collection.sort_values(by="Film ID")

    # Build a label -> Film ID map so selection never depends on re-parsing text
    option_to_id = {}
    for _, row in valid_collection_sorted.iterrows():
      title = row.get("Title", "")
      year_str = safe_year(row.get("Year", ""))
      label = f"{title} ({year_str})" if year_str else f"{title}"
      # De-duplicate identical labels (e.g. two copies of the same title/year)
      base_label, suffix = label, 2
      while base_label in option_to_id:
        base_label = f"{label} [{suffix}]"
        suffix += 1
      option_to_id[base_label] = row.get("Film ID", "")

    selected_option = st.selectbox(
        "Select Movie:", ["-- Select --"] + list(option_to_id.keys())
    )

    if selected_option != "-- Select --":
      selected_id = option_to_id[selected_option]
      movie_row = valid_collection_sorted[
          valid_collection_sorted["Film ID"].astype(str) == str(selected_id)
      ].iloc[0]

      selected_movie = movie_row["Title"]
      current_watched = movie_row.get("Watched", "No")
      current_rating_val = movie_row.get("Rating (1-5)", 3)
      default_star_index = star_index_from_rating(current_rating_val)

      with st.form("update_movie_form"):
        st.markdown(f"**Editing:** {selected_movie}")
        col_a, col_b = st.columns(2)
        with col_a:
          watched_status = st.selectbox(
              "Watched",
              ["Yes", "No"],
              index=0 if is_watched(current_watched) else 1,
          )
        with col_b:
          new_rating_stars = st.selectbox(
              "Stars",
              STAR_OPTIONS,
              index=default_star_index,
          )
        current_date_watched = movie_row.get("Date Watched") if "Date Watched" in movie_row.index else None
        new_date_watched = st.date_input(
            "Date watched (optional)",
            value=pd.to_datetime(current_date_watched).date() if pd.notna(current_date_watched) else datetime.date.today(),
        )

        if st.form_submit_button("💾 Save Changes"):
          if save_collection_update(selected_id, watched_status, new_rating_stars, new_date_watched):
            st.cache_data.clear()
            st.success(f"Saved '{selected_movie}'!")
            if new_rating_stars == STAR_OPTIONS[-1]:
              st.balloons()
            st.rerun()
          else:
            st.error("Couldn't find that film in the sheet to update.")

      film_log = (
          watch_log_df[watch_log_df["Film ID"].astype(str) == str(selected_id)]
          if "Film ID" in watch_log_df.columns else pd.DataFrame()
      )
      if not film_log.empty:
        log_dates = pd.to_datetime(film_log["Date Watched"], errors="coerce").dropna().sort_values()
        if not log_dates.empty:
          times_watched = len(log_dates)
          first_watched = log_dates.min()
          last_watched = log_dates.max()
          years_since_last = (pd.Timestamp.today() - last_watched).days / 365.25
          age_note = f" ({years_since_last:.1f} years ago)" if years_since_last >= 1 else ""
          st.caption(
              f"📼 Watched **{times_watched}** time{'s' if times_watched != 1 else ''} — "
              f"first {first_watched.strftime('%d/%m/%Y')}, last {last_watched.strftime('%d/%m/%Y')}{age_note}"
          )

      with st.expander("📼 Log a rewatch"):
        st.caption("For logging an additional watch without changing today's Watched/Rating fields above.")
        rewatch_date = st.date_input("Date watched", value=datetime.date.today(), key=f"rewatch_date_{selected_id}")
        rewatch_rating_choice = st.selectbox(
            "Rating at this watch", ["(keep current rating)"] + STAR_OPTIONS, key=f"rewatch_rating_{selected_id}"
        )
        if st.button("Log this watch", key=f"log_rewatch_{selected_id}"):
          rating_to_use = (
              STAR_OPTIONS[default_star_index] if rewatch_rating_choice == "(keep current rating)"
              else rewatch_rating_choice
          )
          if save_collection_update(selected_id, "Yes", rating_to_use, rewatch_date):
            st.cache_data.clear()
            st.success(f"Logged a watch of '{selected_movie}' on {rewatch_date.strftime('%d/%m/%Y')}!")
            st.rerun()
          else:
            st.error("Couldn't find that film in the sheet to update.")

      with st.expander("🗑️ Remove this film from my collection"):
        st.warning(f"This permanently deletes '{selected_movie}' from your Collection sheet.")
        confirm_delete = st.checkbox("Yes, I'm sure", key=f"confirm_del_{selected_id}")
        if st.button("Remove permanently", key=f"remove_btn_{selected_id}", disabled=not confirm_delete):
          if delete_from_collection(selected_id):
            st.cache_data.clear()
            st.success(f"Removed '{selected_movie}' from your collection.")
            st.rerun()
          else:
            st.error("Couldn't find that film in the sheet to remove.")

    film_divider()
    st.markdown("**➕ Add a Film**")

    add_tab1, add_tab2 = st.tabs(["Type it in", "📷 Scan barcode"])

    with add_tab1:
      st.markdown("**1. Title & year**")
      mcol1, mcol2 = st.columns([2, 1])
      with mcol1:
        manual_title = st.text_input("Title *", key="manual_title_input")
      with mcol2:
        manual_year = st.number_input(
            "Year", min_value=1900, max_value=2100,
            value=datetime.date.today().year, step=1, key="manual_year_input",
        )

      if manual_title:
        possible_dupes = find_possible_duplicates(manual_title)
        if possible_dupes:
          dupe_list = ", ".join(f"{t} ({safe_year(y)})" for t, y in possible_dupes[:3])
          st.warning(f"⚠️ You might already own this: {dupe_list}. Still fine to add if it's a different edition.")

      if tmdb_configured():
        if st.button("🔎 Look up on TMDb", key="manual_tmdb_btn"):
          if manual_title:
            result = tmdb_lookup_cached(manual_title, manual_year)
            st.session_state["manual_tmdb_result"] = result
            if not result:
              st.warning("No TMDb match — you can still fill in the details manually below.")
          else:
            st.warning("Enter a title first.")
      else:
        st.caption("Add a free TMDb API key to auto-fill genre, director, runtime, and plot here.")

      tmdb_prefill = st.session_state.get("manual_tmdb_result")
      if tmdb_prefill:
        st.success(
            f"Found on TMDb: {tmdb_prefill.get('genre') or '—'} · "
            f"dir. {tmdb_prefill.get('director') or '—'} · "
            f"{tmdb_prefill.get('runtime') or '—'} min"
        )

      st.markdown("**2. Confirm & add**")
      with st.form("manual_add_form", clear_on_submit=True):
        mcol3, mcol4 = st.columns(2)
        with mcol3:
          m_format = st.selectbox("Format", ["Blu-ray", "4K UHD", "DVD", "Other"])
        with mcol4:
          m_runtime = st.number_input(
              "Runtime (min)", min_value=0, max_value=600,
              value=int(tmdb_prefill.get("runtime") or 0) if tmdb_prefill else 0,
              step=1,
          )
        m_genre = st.text_input("Genre", value=(tmdb_prefill.get("genre", "") if tmdb_prefill else ""))
        m_director = st.text_input("Director", value=(tmdb_prefill.get("director", "") if tmdb_prefill else ""))
        m_cast = st.text_input(
            "Cast (comma-separated, top 5)",
            value=(", ".join(tmdb_prefill.get("cast", [])) if tmdb_prefill else ""),
        )
        m_cert_prefill = tmdb_prefill.get("certificate") if tmdb_prefill else None
        m_certificate = st.selectbox(
            "BBFC Rating",
            BBFC_RATINGS,
            index=BBFC_RATINGS.index(m_cert_prefill) if m_cert_prefill in BBFC_RATINGS else 0,
        )
        m_notes = st.text_area(
            "Notes / Plot", value=(tmdb_prefill.get("plot", "") if tmdb_prefill else ""), height=80,
        )

        if st.form_submit_button("💾 Add to Collection"):
          if manual_title:
            add_to_collection({
                "Title": manual_title,
                "Year": manual_year,
                "Format": m_format,
                "Genre": m_genre,
                "Director": m_director,
                "Runtime (min)": m_runtime if m_runtime else None,
                "BBFC Rating": m_certificate if m_certificate else None,
                "Notes": m_notes,
                "Watched": "No",
                **cast_fields_from_string(m_cast),
            }, full_cast=(tmdb_prefill.get("full_cast") if tmdb_prefill else None))
            st.session_state.pop("manual_tmdb_result", None)
            st.cache_data.clear()
            st.success(f"Added '{manual_title}' to your collection!")
            st.rerun()
          else:
            st.warning("Title is required.")

    with add_tab2:
      st.caption(
          "Tap below and choose 'Take Photo' — this opens your phone's real "
          "camera app (which defaults to the rear camera), not a browser "
          "preview. Works best with good lighting and the barcode filling "
          "most of the frame."
      )
      barcode_photo = st.file_uploader(
          "Scan barcode",
          type=["jpg", "jpeg", "png"],
          key="barcode_uploader",
      )

      if barcode_photo is not None:
        image_bytes = barcode_photo.getvalue()
        decoded_code = decode_barcode(image_bytes)
        if not decoded_code:
          st.error("Couldn't read a barcode in that photo — try again with the barcode closer and well-lit.")
        else:
          st.info(f"Barcode: {decoded_code}")
          upc_result = lookup_barcode(decoded_code)

          prefill_title = clean_title_for_search(upc_result.get("title", "")) if upc_result else ""
          prefill_year = upc_result.get("year") if upc_result else None

          # Chain a TMDb lookup off the (cleaned) UPC title to also pull
          # genre/director/runtime/plot, and a more reliable year if TMDb
          # has one -- the UPC database frequently doesn't.
          tmdb_result = None
          if prefill_title and tmdb_configured():
            tmdb_result = tmdb_lookup_cached(prefill_title, prefill_year)
            if tmdb_result and tmdb_result.get("year"):
              prefill_year = tmdb_result["year"]

          possible_dupes = find_possible_duplicates(prefill_title) if prefill_title else []
          if possible_dupes:
            dupe_list = ", ".join(f"{t} ({safe_year(y)})" for t, y in possible_dupes[:3])
            st.warning(f"⚠️ You might already own this: {dupe_list}. Still fine to add if it's a different edition.")

          with st.form("barcode_add_form"):
            if upc_result:
              st.success("Found a match online — check it's right before adding:")
            else:
              st.warning("No online match for this barcode — enter the details manually.")

            if upc_result and not prefill_year:
              st.caption("⚠️ Couldn't confirm a release year — please check it.")
            if tmdb_result:
              st.caption(
                  f"TMDb: {tmdb_result.get('genre') or '—'} · "
                  f"dir. {tmdb_result.get('director') or '—'} · "
                  f"{tmdb_result.get('runtime') or '—'} min"
              )

            b_title = st.text_input("Title *", value=prefill_title)
            bcol1, bcol2 = st.columns(2)
            with bcol1:
              b_year = st.number_input(
                  "Year",
                  min_value=1900,
                  max_value=2100,
                  value=prefill_year if prefill_year else datetime.date.today().year,
                  step=1,
                  key="b_year",
              )
            with bcol2:
              b_format = st.selectbox("Format", ["Blu-ray", "4K UHD", "DVD", "Other"], key="b_format")
            b_genre = st.text_input("Genre", value=(tmdb_result.get("genre", "") if tmdb_result else ""))
            b_director = st.text_input("Director", value=(tmdb_result.get("director", "") if tmdb_result else ""))
            b_cast = st.text_input(
                "Cast (comma-separated, top 5)",
                value=(", ".join(tmdb_result.get("cast", [])) if tmdb_result else ""),
            )
            b_runtime = st.number_input(
                "Runtime (min)", min_value=0, max_value=600,
                value=int(tmdb_result.get("runtime") or 0) if tmdb_result else 0, step=1,
            )
            b_cert_prefill = tmdb_result.get("certificate") if tmdb_result else None
            b_certificate = st.selectbox(
                "BBFC Rating",
                BBFC_RATINGS,
                index=BBFC_RATINGS.index(b_cert_prefill) if b_cert_prefill in BBFC_RATINGS else 0,
            )
            default_notes = f"Barcode: {decoded_code}"
            if tmdb_result and tmdb_result.get("plot"):
              default_notes = f"{tmdb_result['plot']}\n\nBarcode: {decoded_code}"
            b_notes = st.text_area("Notes / Plot", value=default_notes, height=80)

            if st.form_submit_button("💾 Add to Collection"):
              if b_title:
                add_to_collection({
                    "Title": b_title,
                    "Year": b_year,
                    "Format": b_format,
                    "Genre": b_genre,
                    "Director": b_director,
                    "Runtime (min)": b_runtime if b_runtime else None,
                    "BBFC Rating": b_certificate if b_certificate else None,
                    "Notes": b_notes,
                    "Watched": "No",
                    **cast_fields_from_string(b_cast),
                }, full_cast=(tmdb_result.get("full_cast") if tmdb_result else None))
                st.cache_data.clear()
                st.success(f"Added '{b_title}' to your collection!")
                st.rerun()
              else:
                st.warning("Title is required.")

    film_divider()

    with st.expander("🔄 Backfill Genre/Cast/BBFC Rating from TMDb (one-time)"):
      st.caption(
          "Fills in Genre, Director, full Cast, Runtime, Notes, and BBFC "
          "Rating for films that don't have them yet — like ones added "
          "before these features existed. Also expands the Actors/Directors "
          "sheets to include a film's FULL cast (not just the 5 stored in "
          "Collection's own Actor 1-5 columns), so someone who wasn't "
          "top-billed still shows up under Browse by Actor. Never "
          "overwrites anything you've already filled in, and it's safe to "
          "run more than once — films already counted for someone won't be "
          "double-counted."
      )
      if not tmdb_configured():
        st.info("Add a free TMDb API key to your secrets to use this — see the Add tab for setup notes.")
      else:
        st.write(f"This will look up all **{len(valid_collection)}** films in your collection.")
        confirm_backfill = st.checkbox("Yes, look these up and fill in what's missing", key="confirm_backfill")
        if st.button("Run backfill", disabled=not confirm_backfill):
          films_to_process = [
              (row.get("Film ID"), row.get("Title", ""), row.get("Year") if pd.notna(row.get("Year")) else None)
              for _, row in valid_collection.iterrows()
              if row.get("Film ID") and row.get("Title")
          ]
          with st.spinner(f"Looking up {len(films_to_process)} films on TMDb — this can take a couple of minutes..."):
            updated, no_match = backfill_missing_details(films_to_process)
          st.cache_data.clear()
          st.success(f"Processed {len(films_to_process)} films. {updated} had a blank cell filled in. {no_match} had no TMDb match.")
          st.rerun()

    st.markdown("**Your collection**")
    st.dataframe(
        style_format_column(curated_browse_view(valid_collection)),
        use_container_width=True,
        height=350,
        hide_index=True,
    )

  # --- TAB 4: WISHLIST ---
  with app_mode[3]:
    st.subheader("🛒 Wishlist")

    st.markdown("### ➕ Add to Wishlist")
    with st.form("wishlist_form", clear_on_submit=True):
      new_title = st.text_input("Title *")
      new_priority = st.slider("Priority (1-5)", 1, 5, 3)
      new_target_price = st.number_input(
          "Target Price (£)", min_value=0.0, value=15.00, step=0.50
      )
      new_notes = st.text_input("Notes")

      if st.form_submit_button("💾 Add Film"):
        if new_title:
          add_to_wishlist(new_title, new_priority, new_target_price, new_notes)
          st.cache_data.clear()
          st.success(f"Added '{new_title}'!")
          st.rerun()
        else:
          st.warning("Title is required.")

    film_divider()

    if tmdb_configured():
      with st.expander("✨ Because you own these..."):
        rcol1, rcol2 = st.columns([3, 1])
        with rcol1:
          st.caption("Pulled from a handful of random picks across your whole collection.")
        with rcol2:
          refresh_clicked = st.button("🔄 New picks", key="refresh_recs")

        if refresh_clicked or "wishlist_rec_sources" not in st.session_state:
          sample_n = min(5, len(valid_collection))
          if sample_n > 0:
            sampled = valid_collection.sample(n=sample_n)
            st.session_state["wishlist_rec_sources"] = sampled[["Title", "Year"]].to_dict("records")
          else:
            st.session_state["wishlist_rec_sources"] = []

        source_rows = st.session_state.get("wishlist_rec_sources", [])

        if not source_rows:
          st.caption("Add some films to your collection and suggestions will show up here.")
        else:
          owned_titles = set(valid_collection["Title"].dropna().str.strip().str.lower())
          wishlist_titles = set(valid_wishlist["Title"].dropna().str.strip().str.lower()) if not valid_wishlist.empty else set()

          seen = set()
          suggestions = []
          for source_row in source_rows:
            source_lookup = tmdb_lookup_cached(source_row.get("Title", ""), source_row.get("Year"))
            if not source_lookup or not source_lookup.get("id"):
              continue
            for rec in tmdb_recommendations_cached(source_lookup["id"]):
              rec_title = rec.get("title", "")
              key = rec_title.strip().lower()
              if not rec_title or key in seen or key in owned_titles or key in wishlist_titles:
                continue
              seen.add(key)
              suggestions.append(rec)
            if len(suggestions) >= 10:
              break

          if not suggestions:
            st.caption("No new suggestions from these picks — try refreshing for a different set.")
          else:
            for rec in suggestions[:10]:
              rcol1, rcol2 = st.columns([1, 3])
              with rcol1:
                if rec.get("poster_path"):
                  st.image(f"https://image.tmdb.org/t/p/w200{rec['poster_path']}")
              with rcol2:
                rec_label = f"{rec['title']} ({rec['year']})" if rec.get("year") else rec["title"]
                st.markdown(f"**{rec_label}**")
                if st.button("+ Add to Wishlist", key=f"rec_add_{rec['title']}"):
                  add_to_wishlist(rec["title"])
                  st.cache_data.clear()
                  st.success(f"Added '{rec['title']}' to your Wishlist!")
                  st.rerun()

    if not valid_wishlist.empty:
      # Deal alerts: cheapest found price has actually dropped to/below target
      has_price_cols = (
          "Target Price (£)" in valid_wishlist.columns
          and "Cheapest Found (£)" in valid_wishlist.columns
      )
      if has_price_cols:
        deals = valid_wishlist[
            valid_wishlist["Cheapest Found (£)"].notna()
            & valid_wishlist["Target Price (£)"].notna()
            & (valid_wishlist["Cheapest Found (£)"] <= valid_wishlist["Target Price (£)"])
        ]
        if not deals.empty:
          st.markdown("### 🔥 Deal Alerts")
          for _, deal_row in deals.iterrows():
            where = deal_row.get("Where", "")
            where_str = f" at {where}" if pd.notna(where) and where else ""
            st.success(
                f"**{deal_row.get('Title','')}** — found for "
                f"£{deal_row.get('Cheapest Found (£)'):.2f} "
                f"(target was £{deal_row.get('Target Price (£)'):.2f}){where_str}"
            )
          film_divider()

      for idx, row in valid_wishlist.assign(
          _priority_sort=pd.to_numeric(valid_wishlist.get("Priority (1-5)"), errors="coerce").fillna(0)
      ).sort_values(by="_priority_sort", ascending=False).iterrows():
        w_title = row.get("Title", "Untitled")
        w_priority = row.get("Priority (1-5)", "")
        w_target = row.get("Target Price (£)", None)
        w_cheapest = row.get("Cheapest Found (£)", None)

        price_bits = []
        if pd.notna(w_target):
          price_bits.append(f"Target: £{w_target:.2f}")
        if pd.notna(w_cheapest):
          price_bits.append(f"Cheapest found: £{w_cheapest:.2f}")
        price_line = " | ".join(price_bits) if price_bits else "No price tracked yet"

        wcol1, wcol2, wcol3 = st.columns([3, 1, 1])
        with wcol1:
          st.markdown(
              f"**{w_title}**  \n<span style='opacity:0.65;font-size:0.85em;'>"
              f"Priority: {w_priority} | {price_line}</span>",
              unsafe_allow_html=True,
          )
        with wcol2:
          buy_clicked = st.button("✅ Bought", key=f"buy_{idx}")
        with wcol3:
          delete_clicked = st.button("🗑️", key=f"del_{idx}")

        links = retailer_search_links(w_title)
        st.markdown(
            f"<span style='font-size:0.85em;'>🔗 Check price: "
            f"<a href='{links['HMV']}' target='_blank'>HMV</a> · "
            f"<a href='{links['Zavvi']}' target='_blank'>Zavvi</a> · "
            f"<a href='{links['Amazon']}' target='_blank'>Amazon</a> · "
            f"<a href='{links['CEX']}' target='_blank'>CEX</a></span>",
            unsafe_allow_html=True,
        )

        with st.expander("🎚️ Change priority"):
          with st.form(f"priority_form_{idx}", clear_on_submit=False):
            new_priority_value = st.slider(
                "Priority (1-5)", 1, 5,
                value=int(w_priority) if pd.notna(w_priority) and str(w_priority).strip() else 3,
                key=f"priority_slider_{idx}",
            )
            if st.form_submit_button("Save priority"):
              book = openpyxl.load_workbook(FILE_PATH)
              wish_sheet = book["Wishlist"]
              excel_row = idx + 5
              wish_sheet.cell(row=excel_row, column=3, value=new_priority_value)
              book.save(FILE_PATH)
              save_and_sync(FILE_PATH)
              st.cache_data.clear()
              st.success(f"Updated priority for '{w_title}' to {new_priority_value}.")
              st.rerun()

        with st.expander("💷 Log a price check"):
          with st.form(f"price_form_{idx}", clear_on_submit=False):
            pcol1, pcol2 = st.columns(2)
            with pcol1:
              logged_price = st.number_input(
                  "Cheapest price found (£)",
                  min_value=0.0,
                  value=float(w_cheapest) if pd.notna(w_cheapest) else 0.0,
                  step=0.50,
                  key=f"price_input_{idx}",
              )
            with pcol2:
              logged_where = st.text_input(
                  "Where",
                  value=str(row.get("Where", "") or ""),
                  key=f"where_input_{idx}",
              )
            if st.form_submit_button("Save price"):
              book = openpyxl.load_workbook(FILE_PATH)
              wish_sheet = book["Wishlist"]
              excel_row = idx + 5
              wish_sheet.cell(row=excel_row, column=5, value=logged_price)
              wish_sheet.cell(row=excel_row, column=6, value=logged_where)
              wish_sheet.cell(
                  row=excel_row, column=7,
                  value=pd.Timestamp.today().strftime("%Y-%m-%d"),
              )
              book.save(FILE_PATH)
              save_and_sync(FILE_PATH)
              st.cache_data.clear()
              st.success(f"Logged price for '{w_title}'.")
              st.rerun()

        if buy_clicked:
          book = openpyxl.load_workbook(FILE_PATH)
          coll_sheet = book["Collection"]
          wish_sheet = book["Wishlist"]

          next_row = coll_sheet.max_row + 1
          new_film_id = next_film_id(collection_df)
          coll_sheet.cell(row=next_row, column=1, value=new_film_id)
          coll_sheet.cell(row=next_row, column=2, value=w_title)
          coll_sheet.cell(row=next_row, column=7, value="No")  # Watched

          # Remove the row from Wishlist (Excel row = df index + header offset)
          wish_sheet.delete_rows(idx + 5)

          book.save(FILE_PATH)
          save_and_sync(FILE_PATH)
          st.cache_data.clear()
          st.success(f"Moved '{w_title}' to your Collection! 🎉")
          st.rerun()

        if delete_clicked:
          book = openpyxl.load_workbook(FILE_PATH)
          wish_sheet = book["Wishlist"]
          wish_sheet.delete_rows(idx + 5)
          book.save(FILE_PATH)
          save_and_sync(FILE_PATH)
          st.cache_data.clear()
          st.info(f"Removed '{w_title}' from wishlist.")
          st.rerun()
    else:
      st.info("Your wishlist is empty — add something above!")

  # --- TAB 5: STATS ---
  with app_mode[4]:
    st.subheader("📊 Collection Insights")

    st.markdown("### 🎊 Yearly Rewind")

    def build_watch_events(collection_df_in, log_df_in):
      """Combines proper Watch Log entries with a synthetic single entry
      for any film whose Collection 'Date Watched' cache isn't otherwise
      represented in the log -- covers watches saved before the Watch Log
      existed, so old data isn't just dropped."""
      events = []
      if "Film ID" in log_df_in.columns and not log_df_in.empty:
        for _, row in log_df_in.iterrows():
          events.append({
              "Film ID": str(row.get("Film ID", "")),
              "Title": row.get("Title", ""),
              "Date": row.get("Date Watched"),
          })
      logged_pairs = {(e["Film ID"], str(e["Date"])[:10]) for e in events}

      if "Date Watched" in collection_df_in.columns:
        for _, row in collection_df_in.iterrows():
          dw = row.get("Date Watched")
          if pd.isna(dw) or str(dw).strip() == "":
            continue
          fid = str(row.get("Film ID", ""))
          date_str = str(dw)[:10]
          if (fid, date_str) not in logged_pairs:
            events.append({"Film ID": fid, "Title": row.get("Title", ""), "Date": date_str})

      return pd.DataFrame(events)

    watch_events = build_watch_events(valid_collection, watch_log_df)

    if watch_events.empty:
      st.info("Log some watch dates from the Update tab and your yearly rewind will show up here.")
    else:
      watch_events["Date"] = pd.to_datetime(watch_events["Date"], errors="coerce")
      watch_events = watch_events.dropna(subset=["Date"])
      years_available = sorted(watch_events["Date"].dt.year.unique().tolist(), reverse=True)

      if not years_available:
        st.info("No watch dates logged yet — mark a few films watched with a date and check back.")
      else:
        selected_rewind_year = st.selectbox("Year", years_available, key="rewind_year")
        year_events = watch_events[watch_events["Date"].dt.year == selected_rewind_year].copy()

        enrich_cols = [c for c in ["Film ID", "Genre", "Director", "Runtime (min)"] if c in valid_collection.columns]
        if "Film ID" in enrich_cols and len(enrich_cols) > 1:
          enrich_df = valid_collection[enrich_cols].copy()
          enrich_df["Film ID"] = enrich_df["Film ID"].astype(str)
          year_events = year_events.merge(enrich_df, on="Film ID", how="left")

        total_sessions = len(year_events)
        unique_films = year_events["Title"].nunique()

        top_genre = None
        if "Genre" in year_events.columns:
          genre_counts = split_multi_value_counts(year_events["Genre"])
          top_genre = genre_counts.index[0] if not genre_counts.empty else None

        top_director = None
        if "Director" in year_events.columns:
          director_counts = split_multi_value_counts(year_events["Director"])
          top_director = director_counts.index[0] if not director_counts.empty else None

        busiest_month = None
        if total_sessions:
          month_counts = year_events["Date"].dt.strftime("%B").value_counts()
          busiest_month = month_counts.index[0] if not month_counts.empty else None

        longest_title = shortest_title = None
        total_hours = None
        if "Runtime (min)" in year_events.columns:
          runtimes = pd.to_numeric(year_events["Runtime (min)"], errors="coerce").dropna()
          if not runtimes.empty:
            longest_idx = runtimes.idxmax()
            shortest_idx = runtimes.idxmin()
            longest_title = f"{year_events.loc[longest_idx, 'Title']} ({int(runtimes.loc[longest_idx])} min)"
            shortest_title = f"{year_events.loc[shortest_idx, 'Title']} ({int(runtimes.loc[shortest_idx])} min)"
            total_hours = runtimes.sum() / 60

        rwcol1, rwcol2, rwcol3 = st.columns(3)
        with rwcol1:
          st.metric("Watch Sessions", total_sessions)
        with rwcol2:
          st.metric("Unique Films", unique_films)
        with rwcol3:
          st.metric("Hours Watched", f"{total_hours:.1f}" if total_hours else "—")

        if top_genre:
          st.markdown(f"**Top Genre:** {top_genre}")
        if top_director:
          st.markdown(f"**Top Director:** {top_director}")
        if busiest_month:
          st.markdown(f"**Busiest Month:** {busiest_month}")
        if longest_title:
          st.markdown(f"**Longest Watch:** {longest_title}")
        if shortest_title:
          st.markdown(f"**Shortest Watch:** {shortest_title}")

    film_divider()

    if "Genre" in valid_collection.columns:
      st.markdown("**By Genre**")
      genre_counts = split_multi_value_counts(valid_collection["Genre"]).head(15)
      st.bar_chart(genre_counts)

    if "Year" in valid_collection.columns:
      st.markdown("**By Decade**")
      decades = (
          pd.to_numeric(valid_collection["Year"], errors="coerce")
          .dropna()
          .apply(lambda y: f"{int(y) // 10 * 10}s")
      )
      st.bar_chart(decades.value_counts().sort_index())

    if "Director" in valid_collection.columns:
      st.markdown("**Top Directors**")
      top_directors = split_multi_value_counts(valid_collection["Director"]).head(10)
      st.bar_chart(top_directors)

    if "Watched" in valid_collection.columns:
      watched_count = valid_collection["Watched"].apply(is_watched).sum()
      unwatched_count = total_collection - watched_count
      st.markdown("**Watched vs Unwatched**")
      st.bar_chart(pd.Series(
          {"Watched": watched_count, "Unwatched": unwatched_count}
      ))

    if "Date Watched" in valid_collection.columns:
      last_watched_dates = pd.to_datetime(valid_collection["Date Watched"], errors="coerce")
      forgotten = valid_collection.copy()
      forgotten["_last_watched"] = last_watched_dates
      forgotten = forgotten.dropna(subset=["_last_watched"])
      forgotten["_years_ago"] = (pd.Timestamp.today() - forgotten["_last_watched"]).dt.days / 365.25
      forgotten = forgotten[forgotten["_years_ago"] >= 3].sort_values(by="_years_ago", ascending=False)

      if not forgotten.empty:
        film_divider()
        st.markdown("**🕰️ Forgotten Favorites** (not watched in 3+ years)")
        for _, frow in forgotten.head(10).iterrows():
          st.markdown(
              f"- **{frow['Title']}** ({safe_year(frow.get('Year'))}) — "
              f"last watched {frow['_last_watched'].strftime('%d/%m/%Y')} "
              f"({frow['_years_ago']:.1f} years ago)"
          )

  # --- TAB 6: EXTRAS ---
  with app_mode[5]:
    with st.expander("💾 Backup & Export"):
      st.caption(
          "Downloads the exact spreadsheet the app is currently using -- "
          "same data as what's synced to GitHub, just handy to grab a copy "
          "any time without needing to go find it in the repo."
      )
      try:
        with open(FILE_PATH, "rb") as f:
          file_bytes = f.read()
        backup_filename = f"Blu-ray_Collection_{datetime.date.today().strftime('%Y-%m-%d')}.xlsx"
        st.download_button(
            "📥 Download Current Database",
            data=file_bytes,
            file_name=backup_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
      except FileNotFoundError:
        st.error("Couldn't find the spreadsheet file to download.")

      if github_sync_configured():
        st.caption(
            "Deletes and bulk backfills also save a timestamped safety copy "
            "to a `backups/` folder in your GitHub repo automatically."
        )

    st.subheader("🏆 Franchises, Awards & Ratings")
    st.caption(
        "Franchise % Complete and Consensus scores are Excel formulas. "
        "If you've just made an edit and they look blank, open the file in "
        "Excel once to refresh them (this app tries to do it automatically "
        "if LibreOffice is installed on your machine)."
    )
    sub_tab1, sub_tab2, sub_tab3, sub_tab4, sub_tab5, sub_tab6 = st.tabs(
        ["Franchises", "Awards", "Ratings", "People", "Milestones", "Recently Watched"]
    )
    with sub_tab1:
      if (
          "Collection" in valid_collection.columns
          and "Collection" in franchise_df.columns
          and "Target Films" in franchise_df.columns
      ):
        owned_by_set = valid_collection["Collection"].dropna().value_counts()
        sets_progress = franchise_df.dropna(subset=["Collection"]).copy()
        sets_progress["Owned"] = (
            sets_progress["Collection"].map(owned_by_set).fillna(0).astype(int)
        )
        sets_progress["Target"] = pd.to_numeric(
            sets_progress["Target Films"], errors="coerce"
        ).fillna(sets_progress["Owned"])
        sets_progress["Missing"] = (
            (sets_progress["Target"] - sets_progress["Owned"]).clip(lower=0).astype(int)
        )
        sets_progress = sets_progress.sort_values(
            by="Missing", ascending=False
        )

        st.markdown("**📦 Box Set Completion**")
        for _, srow in sets_progress.iterrows():
          owned = int(srow["Owned"])
          target = int(srow["Target"])
          missing = int(srow["Missing"])
          pct = min(1.0, owned / target) if target > 0 else 0.0
          label = f"{srow['Collection']} — {owned}/{target}"
          if missing == 0 and target > 0:
            label += " ✅ Complete!"
          else:
            label += f" ({missing} to go)"
          st.markdown(label)
          st.progress(pct)
        film_divider()

      st.dataframe(franchise_df, use_container_width=True, height=300)
    with sub_tab2:
      if {"Title", "Nominations"}.issubset(awards_df.columns):
        nominated = awards_df[
            pd.to_numeric(awards_df["Nominations"], errors="coerce").fillna(0) > 0
        ].copy()
        nominated["Wins"] = pd.to_numeric(nominated.get("Wins", 0), errors="coerce").fillna(0).astype(int)
        nominated["Nominations"] = pd.to_numeric(nominated["Nominations"], errors="coerce").fillna(0).astype(int)
        nominated = nominated.sort_values(by=["Wins", "Nominations"], ascending=False)

        if nominated.empty:
          st.info("None of your films show any Oscar nominations in this sheet.")
        else:
          st.caption(f"{len(nominated)} films with at least one Oscar nomination.")
          for _, arow in nominated.iterrows():
            title = arow.get("Title", "Unknown")
            year_str = safe_year(arow.get("Year", ""))
            wins = int(arow.get("Wins", 0))
            noms = int(arow.get("Nominations", 0))
            categories = arow.get("Winning Categories") if wins > 0 else arow.get("All Nominated Categories")
            categories = categories if pd.notna(categories) else ""

            if wins > 0:
              headline = f"🏆 **{title}** ({year_str}) — {wins} win{'s' if wins != 1 else ''}, {noms} nomination{'s' if noms != 1 else ''}"
            else:
              headline = f"🎖️ **{title}** ({year_str}) — {noms} nomination{'s' if noms != 1 else ''}"
            st.markdown(headline)
            if categories:
              st.caption(categories)
      else:
        st.dataframe(awards_df, use_container_width=True, height=300)
    with sub_tab3:
      if "Rating (1-5)" in valid_collection.columns:
        rated = valid_collection.copy()
        rated["_stars"] = rated["Rating (1-5)"].apply(stars_to_number)
        rated = rated[rated["_stars"].notna()].sort_values(by="_stars", ascending=False)

        if rated.empty:
          st.info("No films rated yet — rate some from the Update tab and they'll show up here.")
        else:
          st.caption(f"{len(rated)} rated film{'s' if len(rated) != 1 else ''}, highest first.")
          for _, rrow in rated.iterrows():
            title = rrow.get("Title", "Unknown")
            year_str = safe_year(rrow.get("Year", ""))
            stars = rating_stars_display(rrow.get("Rating (1-5)"))
            st.markdown(f"<span class='stars-display'>{stars}</span> — **{title}** ({year_str})", unsafe_allow_html=True)
      else:
        st.dataframe(ratings_df, use_container_width=True, height=300)

    with sub_tab4:
      person_type = st.radio("Browse by", ["Director", "Actor"], horizontal=True, key="people_browse_type")
      people_df = directors_df if person_type == "Director" else actors_df
      name_col = "Director" if person_type == "Director" else "Actor"

      if name_col not in people_df.columns or "Films Owned" not in people_df.columns:
        st.info(f"No {person_type.lower()} data found in this sheet.")
      else:
        people_sorted = people_df.dropna(subset=[name_col]).sort_values(by="Films Owned", ascending=False)
        options = [
            f"{row[name_col]} ({int(row['Films Owned'])} films)"
            for _, row in people_sorted.iterrows()
        ]
        name_lookup = dict(zip(options, people_sorted[name_col]))

        selected_person_option = st.selectbox(f"Select a {person_type.lower()}:", ["-- Select --"] + options)

        if selected_person_option != "-- Select --":
          selected_name = name_lookup[selected_person_option]
          person_row = people_sorted[people_sorted[name_col] == selected_name].iloc[0]

          pcol1, pcol2, pcol3 = st.columns(3)
          with pcol1:
            st.metric("Films Owned", int(person_row.get("Films Owned", 0)))
          with pcol2:
            st.metric("Earliest", safe_year(person_row.get("Earliest Film")))
          with pcol3:
            st.metric("Latest", safe_year(person_row.get("Latest Film")))

          film_ids_raw = person_row.get("Film IDs", "")
          film_ids = [f.strip() for f in str(film_ids_raw).split(",") if f.strip()] if pd.notna(film_ids_raw) else []

          if film_ids and "Film ID" in valid_collection.columns:
            person_films = valid_collection[valid_collection["Film ID"].astype(str).isin(film_ids)]
            st.dataframe(
                style_format_column(curated_browse_view(person_films)),
                use_container_width=True,
                hide_index=True,
            )
          else:
            films_text = person_row.get("Films in Collection", "")
            if pd.notna(films_text):
              st.markdown(films_text)


    with sub_tab5:
      st.subheader("🏅 Milestones & Achievements")

      milestone_thresholds = [25, 50, 100, 150, 200, 250, 300, 350, 400, 450, 500]

      def render_milestone(count, label, emoji):
        reached = [t for t in milestone_thresholds if count >= t]
        next_target = next((t for t in milestone_thresholds if count < t), None)
        headline = f"{emoji} **{label}: {count}**"
        if reached:
          headline += f" — {reached[-1]}+ unlocked 🎉"
        st.markdown(headline)
        if next_target:
          st.progress(count / next_target)
          st.caption(f"{next_target - count} more to reach {next_target}")

      render_milestone(total_collection, "Films Owned", "🎬")
      watched_total = int(valid_collection["Watched"].apply(is_watched).sum()) if "Watched" in valid_collection.columns else 0
      render_milestone(watched_total, "Films Watched", "✅")

      if "Best Picture" in valid_collection.columns:
        bp_nominated = (valid_collection["Best Picture"] == "Nominated").sum()
        bp_winner = (valid_collection["Best Picture"] == "Winner").sum()
        bp_total = int(bp_nominated + bp_winner)
        if bp_total:
          render_milestone(bp_total, "Best Picture Nominees Owned", "🏆")
          if bp_winner:
            st.caption(f"({int(bp_winner)} of those are Best Picture winners)")

      film_divider()

      if {"Collection", "Target Films"}.issubset(franchise_df.columns) and "Collection" in valid_collection.columns:
        owned_by_set_ms = valid_collection["Collection"].dropna().value_counts()
        fr_ms = franchise_df.dropna(subset=["Collection"]).copy()
        fr_ms["Owned"] = fr_ms["Collection"].map(owned_by_set_ms).fillna(0).astype(int)
        fr_ms["Target"] = pd.to_numeric(fr_ms["Target Films"], errors="coerce").fillna(fr_ms["Owned"])
        completed_sets = fr_ms[(fr_ms["Target"] > 0) & (fr_ms["Owned"] >= fr_ms["Target"])]

        if not completed_sets.empty:
          st.markdown("**📦 Completed Box Sets**")
          for _, crow in completed_sets.iterrows():
            st.markdown(f"✅ All {int(crow['Target'])} **{crow['Collection']}** films owned!")
          film_divider()

      if "Director" in valid_collection.columns:
        director_counts = split_multi_value_counts(valid_collection["Director"])
        big_directors = director_counts[director_counts >= 5]
        if not big_directors.empty:
          st.markdown("**🎥 Director Milestones**")
          for dname, dcount in big_directors.head(10).items():
            st.markdown(f"🎬 {int(dcount)} **{dname}** films owned")

    with sub_tab6:
      st.subheader("🕓 Recently Watched")
      if "Date Watched" not in valid_collection.columns:
        st.info("Log some watch dates from the Collection tab and your recently watched films will show up here.")
      else:
        recent = valid_collection.copy()
        recent["_last_watched"] = pd.to_datetime(recent["Date Watched"], errors="coerce")
        recent = recent.dropna(subset=["_last_watched"]).sort_values(by="_last_watched", ascending=False).head(5)

        if recent.empty:
          st.info("No watch dates logged yet — mark a film watched with a date from the Collection tab.")
        else:
          for _, rrow in recent.iterrows():
            title = rrow.get("Title", "Unknown")
            year_str = safe_year(rrow.get("Year"))
            watched_date = rrow["_last_watched"].strftime("%d/%m/%Y")

            stars_line = f"Watched {watched_date}"
            if "Rating (1-5)" in rrow.index:
              stars = rating_stars_display(rrow.get("Rating (1-5)"))
              if stars != "Not rated yet":
                stars_line += f" · <span class='stars-display'>{stars}</span>"

            pcol1, pcol2 = st.columns([1, 3])
            with pcol1:
              poster_path = None
              if tmdb_configured():
                poster_lookup = tmdb_lookup_cached(title, rrow.get("Year"))
                if poster_lookup:
                  poster_path = poster_lookup.get("poster_path")
              if poster_path:
                st.image(f"https://image.tmdb.org/t/p/w200{poster_path}")
              else:
                st.caption("(no poster found)")
            with pcol2:
              st.markdown(f"🎬 **{title}** ({year_str})")
              st.markdown(f"<span style='opacity:0.7; font-size:0.9em;'>{stars_line}</span>", unsafe_allow_html=True)

            film_divider()

  # --- TAB 7: ON THIS DAY ---
  with app_mode[6]:
    st.subheader("📅 On This Day")
    today = datetime.date.today()
    st.caption(f"Films in your collection first released on {today.strftime('%B %d')}, across the years.")

    if not tmdb_configured():
      st.info(
          "This needs a free TMDb API key in your app's secrets "
          "(the same [tmdb] section used for auto-fill on Add) — "
          "see the Extras tab caption for the general secrets pattern."
      )
    else:
      with st.spinner("Checking release dates across your collection — cached after the first time, so this gets fast."):
        titles_years = [
            (row.get("Title", ""), row.get("Year") if pd.notna(row.get("Year")) else None)
            for _, row in valid_collection.iterrows()
            if row.get("Title")
        ]

        def _check(title_year):
          title, year = title_year
          result = tmdb_release_date_cached(title, year)
          if not result or not result.get("release_date"):
            return None
          try:
            d = datetime.datetime.strptime(result["release_date"], "%Y-%m-%d").date()
          except ValueError:
            return None
          if d.month == today.month and d.day == today.day:
            return (title, d.year, result.get("poster_path"))
          return None

        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
          results = list(executor.map(_check, titles_years))

      matches = sorted((r for r in results if r), key=lambda x: x[1])

      if matches:
        for title, rel_year, poster_path in matches:
          pcol1, pcol2 = st.columns([1, 3])
          with pcol1:
            if poster_path:
              st.image(f"https://image.tmdb.org/t/p/w200{poster_path}")
          with pcol2:
            st.markdown(f"🎬 **{title}**")
            st.caption(f"Released {today.strftime('%B %d')}, {rel_year}")
      else:
        st.info("Nothing in your collection was first released on this date — check back tomorrow!")

except FileNotFoundError:
  st.error(
      f"Couldn't find '{FILE_PATH}'. Make sure the spreadsheet is in the "
      "same folder as this app."
  )
except Exception as e:
  st.error(f"Error: {e}")
